import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
import plotly.express as px
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
import base64
import os
import requests as req
from openpyxl import load_workbook
import re

from pathlib import Path

import tempfile
import base64
import requests as req

FILE_PATH = os.path.join(
    tempfile.gettempdir(),
    "World_Cup_2026_Comp.xlsx"
)

def download_workbook():

    share_id = (
        "u!" +
        base64.b64encode(
            SHAREPOINT_FILE.encode()
        ).decode()
        .rstrip("=")
        .replace("/", "_")
        .replace("+", "-")
    )

    token = os.getenv("GRAPH_ACCESS_TOKEN")

    url = (
        f"https://graph.microsoft.com/v1.0/"
        f"shares/{share_id}/driveItem/content"
    )

    r = req.get(
        url,
        headers={
            "Authorization": f"Bearer {token}"
        },
        timeout=30
    )

    r.raise_for_status()

    with open(FILE_PATH, "wb") as f:
        f.write(r.content)



def _load_img(path, mime):
    try:
        with open(path, "rb") as f:
            return f"data:{mime};base64," + base64.b64encode(f.read()).decode()
    except:
        return ""

BASE_DIR = Path(__file__).resolve().parent


def _load_img(path, mime):
    try:
        with open(path, "rb") as f:
            return f"data:{mime};base64," + base64.b64encode(f.read()).decode()
    except Exception:
        return ""


FIFA_LOGO    = _load_img(str(BASE_DIR / "FIFA_Logo_2026.webp"), "image/webp")
CONCEPT_LOGO = _load_img(str(BASE_DIR / "Concept Logo World Cup 2026.jpg"), "image/jpeg")

st.set_page_config(
    page_title="FIFA 2026 Dashboard",
    page_icon="⚽",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# =========================
# ✅ FIFA DARK THEME STYLE
# =========================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Roboto+Condensed:wght@300;400;700&display=swap');

/* ── Global background + host-nation ambient glow ── */
.stApp, html, body { background-color: #07091e; color: #e0e4f4; }
.stApp::before {
    content: '';
    position: fixed;
    inset: 0;
    background:
        radial-gradient(ellipse 55% 90% at 10% 50%, rgba(0,60,140,0.09) 0%, transparent 70%),
        radial-gradient(ellipse 55% 90% at 90% 50%, rgba(190,20,20,0.06) 0%, transparent 70%),
        radial-gradient(ellipse 80% 45% at 50% 105%, rgba(0,120,50,0.06) 0%, transparent 60%);
    pointer-events: none;
    z-index: 0;
}
/* ── Diagonal line tech texture ── */
.stApp::after {
    content: '';
    position: fixed;
    inset: 0;
    background-image: repeating-linear-gradient(
        -55deg,
        transparent 0px, transparent 28px,
        rgba(212,175,55,0.016) 28px, rgba(212,175,55,0.016) 29px
    );
    pointer-events: none;
    z-index: 0;
}
/* ── Full-width layout (no sidebar) ── */
.block-container {
    padding-top: 1.2rem !important;
    padding-left: 2.5rem !important;
    padding-right: 2.5rem !important;
    max-width: 100% !important;
}

/* ── FIFA header banner ── */
.fifa-banner {
    background:
        linear-gradient(160deg, rgba(0,23,74,0.94) 0%, rgba(0,13,46,0.91) 45%, rgba(0,13,46,0.91) 55%, rgba(0,26,82,0.94) 100%),
        url('https://images.unsplash.com/photo-1522778119026-d647f0596c20?w=1400&q=30') center/cover no-repeat;
    border: 1px solid #1a2f6b;
    border-bottom: 3px solid #d4af37;
    border-radius: 16px;
    padding: 20px 28px 18px;
    text-align: center;
    margin-bottom: 18px;
    position: relative;
    overflow: hidden;
}
.fifa-banner::before {
    content: '2026';
    font-family: 'Bebas Neue', Impact, sans-serif;
    position: absolute;
    font-size: 150px;
    letter-spacing: 24px;
    line-height: 1;
    top: 50%;
    left: 50%;
    transform: translate(-50%, -50%);
    color: transparent;
    -webkit-text-stroke: 1px rgba(212,175,55,0.11);
    pointer-events: none;
    z-index: 0;
    user-select: none;
    white-space: nowrap;
}
/* Sheen sweep across banner */
.fifa-banner::after {
    content: '';
    position: absolute;
    top: 0; left: -120%; width: 70%; height: 100%;
    background: linear-gradient(105deg, transparent 0%, rgba(255,255,255,0.025) 50%, transparent 100%);
    animation: bannerSheen 7s ease-in-out infinite;
    pointer-events: none;
    z-index: 0;
}
.fifa-banner > * { position: relative; z-index: 1; }
.fifa-banner-logo {
    font-size: 30px;
    font-weight: 900;
    letter-spacing: 6px;
    color: #4a90d9;
    font-family: Arial Black, sans-serif;
    text-transform: uppercase;
    margin-bottom: 2px;
}
.fifa-title {
    font-family: 'Bebas Neue', Impact, sans-serif;
    color: #ffffff;
    font-size: 50px;
    letter-spacing: 5px;
    margin: 0;
    line-height: 1;
    text-shadow: 0 0 50px rgba(255,255,255,0.25), 0 2px 14px rgba(0,0,0,0.95);
    animation: titleBreath 5s ease-in-out infinite;
}
.fifa-hosts {
    margin-top: 6px;
    font-size: 12px;
    letter-spacing: 2px;
    font-family: 'Roboto Condensed', Arial, sans-serif;
    font-weight: 700;
    text-transform: uppercase;
}
.fifa-hosts .canada { color: #ff6666; }
.fifa-hosts .mexico { color: #55cc77; }
.fifa-hosts .usa    { color: #66aaff; }
.fifa-hosts .sep    { color: #2a3a6a; margin: 0 10px; }
.fifa-dates {
    color: rgba(255,255,255,0.6);
    font-size: 11px;
    letter-spacing: 4px;
    text-transform: uppercase;
    margin-top: 6px;
}

/* ── Banner flex layout (logo | title | logo) ── */
.banner-flex { display:flex; align-items:center; justify-content:center; gap:12px; }
.banner-center { flex:0 1 auto; text-align:center; }
.banner-logo-left  { flex:0 0 auto; text-align:center; }
.banner-logo-right { flex:0 0 auto; text-align:center; }

/* ── Section titles ── */
.section-title {
    font-family: 'Bebas Neue', Impact, sans-serif;
    color: #ffffff;
    font-size: 20px;
    letter-spacing: 3px;
    text-transform: uppercase;
    border-left: 4px solid #d4af37;
    padding-left: 12px;
    margin: 30px 0 14px;
    animation: glowPulse 3s ease-in-out infinite, fadeInUp 0.5s ease both;
}

/* ── Match card ── */
.match-card {
    background: linear-gradient(135deg, #0c1132 0%, #111840 100%);
    border: 1px solid #1c2558;
    border-radius: 14px;
    margin-bottom: 9px;
    overflow: hidden;
    transition: border-color 0.25s, box-shadow 0.25s, transform 0.22s;
    animation: fadeInUp 0.4s ease both;
}
.match-card:hover { border-color: #d4af37; box-shadow: 0 4px 26px rgba(212,175,55,0.14); transform: translateY(-2px); }
.match-card.live {
    border: 2px solid #ff4444;
    background: linear-gradient(135deg, #1b0808 0%, #220e0e 100%);
    box-shadow: 0 0 24px rgba(255,60,60,0.3);
}
.match-card.finished { border-color: #162a16; background: linear-gradient(135deg, #0a130a 0%, #0e1a0e 100%); }

/* ── Match inner layout ── */
.match-inner {
    display: flex;
    align-items: center;
    padding: 12px 18px;
    gap: 6px;
}
/* Fixed-width team blocks so goalscorer columns get the remaining space */
.team-left  { display: flex; align-items: center; gap: 8px; flex-shrink: 0; width: 150px; }
.team-right { display: flex; align-items: center; gap: 8px; flex-shrink: 0; width: 150px; justify-content: flex-end; }
.team-name {
    font-family: 'Roboto Condensed', Arial, sans-serif;
    font-weight: 700;
    font-size: 15px;
    color: #c0c8e8;
    transition: color 0.2s;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}
.team-name.winner { color: #ffd700; text-shadow: 0 0 14px rgba(255,215,0,0.45); }
.team-name.loser  { color: #3a4560; }
.flag-img {
    width: 32px; height: 22px;
    object-fit: cover;
    border-radius: 2px;
    border: 1px solid rgba(255,255,255,0.18);
    flex-shrink: 0;
}
.flag-blank { width: 32px; height: 22px; display: inline-block; flex-shrink: 0; }

/* Goalscorer columns — inline in the same row, between team and score */
.gs-col {
    flex: 1;
    min-width: 0;
    display: flex;
    flex-direction: column;
    gap: 2px;
    justify-content: center;
}
.gs-col.right { align-items: flex-end; }
.gs-entry {
    font-size: 11px;
    color: #7888a8;
    font-family: 'Roboto Condensed', Arial, sans-serif;
    display: flex;
    align-items: center;
    gap: 3px;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
    max-width: 100%;
}
.gs-col.right .gs-entry { flex-direction: row-reverse; }
.gs-entry .gs-ball  { font-size: 10px; flex-shrink: 0; }
.gs-entry .gs-name  { overflow: hidden; text-overflow: ellipsis; min-width: 0; }
.gs-entry .gs-min   { color: #d4af37; font-size: 10px; font-weight: 700; flex-shrink: 0; }

/* ── VS / Score centre ── */
.vs-center { width: 160px; flex-shrink: 0; text-align: center; }
.vs-text { font-family: 'Bebas Neue', Impact, sans-serif; font-size: 20px; color: #d4af37; line-height: 1; }
.result-win {
    font-family: 'Bebas Neue', Impact, sans-serif;
    font-size: 11px; color: #80e8a8; line-height: 1.35; letter-spacing: 1px;
}
.result-win.score-line {
    font-size: 28px; letter-spacing: 4px; color: #80e8a8;
    text-shadow: 0 0 18px rgba(128,232,168,0.35);
}
.result-draw {
    font-family: 'Bebas Neue', Impact, sans-serif;
    font-size: 28px; color: #90c8ff; letter-spacing: 4px;
    line-height: 1.2; text-align: center;
    text-shadow: 0 0 14px rgba(144,200,255,0.3);
}
.live-pulse {
    font-family: 'Bebas Neue', Impact, sans-serif;
    font-size: 18px; color: #ff4444; letter-spacing: 2px;
    animation: blinker 1.2s linear infinite;
}

/* ── Match meta (rightmost) ── */
.match-meta { flex-shrink: 0; min-width: 175px; text-align: right; }
.match-date { font-size: 13px; font-weight: 600; color: #b8c4e8; }

/* ── Status badges ── */
.badge {
    display: inline-block;
    padding: 3px 10px;
    border-radius: 20px;
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 1px;
    margin-top: 5px;
}
.badge-live     { background:#ff4444; color:#fff; animation: blinker 1.2s linear infinite, pulseRing 2s ease-out infinite; }
.badge-upcoming { background:#1a4f82; color:#90c8ff; }
.badge-finished { background:#1a5e30; color:#80e8a8; }
.badge-tbc      { background:#333; color:#aaa; }
@keyframes blinker { 50% { opacity: 0.25; } }

/* ── All animation keyframes ── */
@keyframes titleBreath {
    0%, 100% { text-shadow: 0 0 30px rgba(255,255,255,0.18), 0 2px 14px rgba(0,0,0,0.95); }
    50%       { text-shadow: 0 0 80px rgba(255,255,255,0.5), 0 0 30px rgba(180,210,255,0.3), 0 2px 14px rgba(0,0,0,0.95); }
}
@keyframes bannerSheen {
    0%    { left: -120%; }
    40%   { left: 130%;  }
    100%  { left: 130%;  }
}
@keyframes fadeInUp {
    from { opacity: 0; transform: translateY(18px); }
    to   { opacity: 1; transform: translateY(0);    }
}
@keyframes glowPulse {
    0%, 100% { border-left-color: #d4af37; }
    50%       { border-left-color: #ffe97a; filter: drop-shadow(-2px 0 10px rgba(212,175,55,0.55)); }
}
@keyframes floatUp {
    0%, 100% { transform: translateY(0px);  }
    50%       { transform: translateY(-4px); }
}
@keyframes pulseRing {
    0%   { box-shadow: 0 0 0 0   rgba(255,68,68,0.8); }
    70%  { box-shadow: 0 0 0 10px rgba(255,68,68,0);   }
    100% { box-shadow: 0 0 0 0   rgba(255,68,68,0);   }
}
@keyframes barFill {
    from { width: 0; }
}
@keyframes rotateBall {
    from { transform: rotate(0deg); }
    to   { transform: rotate(360deg); }
}

/* ── HTML Leaderboard rows ── */
.lb-container { width: 100%; padding-bottom: 4px; }
.lb-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 8px; }
.lb-col  { display: flex; flex-direction: column; }
.lb-row {
    display: flex; align-items: center; gap: 8px;
    padding: 7px 10px;
    background: linear-gradient(135deg, #0c1132, #111840);
    border: 1px solid #1c2558;
    border-left: 3px solid var(--pc, #1c2558);
    border-radius: 8px; margin-bottom: 4px;
    transition: transform 0.18s, box-shadow 0.18s;
    animation: fadeInUp 0.35s ease both;
}
.lb-row:hover { transform: translateX(4px); box-shadow: -4px 0 18px rgba(0,0,0,0.5); }
.lb-row.rank-1 { border-left-color: #d4af37; background: linear-gradient(135deg, #1a1408 0%, #241c08 100%); }
.lb-row.rank-2 { border-left-color: #909090; background: linear-gradient(135deg, #131313 0%, #1c1c1e 100%); }
.lb-row.rank-3 { border-left-color: #a0522d; background: linear-gradient(135deg, #18100a 0%, #20130b 100%); }
.lb-num { font-family: 'Bebas Neue', Impact, sans-serif; font-size: 15px; min-width: 24px; text-align: center; flex-shrink: 0; }
.lb-num.p1 { color: #d4af37; } .lb-num.p2 { color: #a8a8a8; } .lb-num.p3 { color: #cd7f32; } .lb-num.pn { color: #2a3555; font-size: 12px; }
.lb-avatar { width: 26px; height: 26px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-family: 'Bebas Neue', Impact, sans-serif; font-size: 10px; color: #07091e; flex-shrink: 0; font-weight: 900; }
.lb-name { font-family: 'Roboto Condensed', Arial, sans-serif; font-weight: 700; font-size: 12px; color: #d8ddf0; flex: 1; min-width: 50px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
.lb-bar-wrap { flex: 2; height: 3px; background: #0a0e20; border-radius: 3px; overflow: hidden; }
.lb-bar { height: 100%; border-radius: 3px; animation: barFill 1.2s ease-out both; }
.lb-score { font-family: 'Bebas Neue', Impact, sans-serif; font-size: 16px; min-width: 44px; text-align: right; color: #c8d0f0; flex-shrink: 0; }
.lb-score sup { font-size: 9px; color: #334560; font-family: 'Roboto Condensed'; vertical-align: super; }
.lb-acc { font-size: 9px; min-width: 34px; text-align: right; color: #445070; font-family: 'Roboto Condensed'; flex-shrink: 0; }

/* ── Football pitch section divider ── */
.pitch-divider {
    width: 100%; height: 54px;
    margin: 26px 0 6px;
    position: relative; overflow: hidden;
    border-radius: 10px;
    border: 1px solid rgba(30,80,30,0.6);
    background:
        linear-gradient(rgba(0,0,0,0.42), rgba(0,0,0,0.42)),
        url('https://images.unsplash.com/photo-1431324155629-1a6deb1dec8d?w=1000&q=30') center/cover no-repeat;
    display: flex; align-items: center; justify-content: center; gap: 14px;
}
.pitch-divider::before {
    content: '';
    position: absolute; inset: 0;
    background-image: linear-gradient(rgba(255,255,255,0.04) 1px, transparent 1px),
                      linear-gradient(90deg, rgba(255,255,255,0.04) 1px, transparent 1px);
    background-size: 44px 44px;
}
.pitch-ball { position: relative; z-index: 1; font-size: 24px; animation: rotateBall 4s linear infinite; }
.pitch-label {
    position: relative; z-index: 1;
    font-family: 'Bebas Neue', Impact, sans-serif;
    font-size: 22px; letter-spacing: 7px;
    color: rgba(255,255,255,0.92);
    text-shadow: 0 1px 10px rgba(0,0,0,0.95);
}

.stat-row { display: flex; gap: 14px; margin: 16px 0 22px; flex-wrap: wrap; }
.stat-card {
    flex: 1;
    min-width: 130px;
    background: linear-gradient(135deg, #0c1132, #111840);
    border: 1px solid #1c2558;
    border-radius: 14px;
    padding: 20px 14px 16px;
    text-align: center;
    position: relative;
    overflow: hidden;
    animation: fadeInUp 0.5s ease both, floatUp 4s ease-in-out infinite 0.6s;
    transition: transform 0.25s, box-shadow 0.25s, border-color 0.25s;
}
.stat-card:hover { border-color: #d4af37; transform: translateY(-5px); box-shadow: 0 10px 32px rgba(212,175,55,0.18); }
.stat-card::before {
    content: attr(data-icon);
    position: absolute;
    font-size: 80px;
    top: 50%;
    left: 50%;
    transform: translate(-50%, -50%);
    opacity: 0.055;
    pointer-events: none;
    line-height: 1;
}
.stat-label { font-size: 10px; font-weight: 700; letter-spacing: 2px; text-transform: uppercase; color: #7888b8; margin-bottom: 4px; }
.stat-value { font-family: 'Bebas Neue', Impact, sans-serif; font-size: 50px; line-height: 1.05; color: #e0e4f4; }
.stat-value.gold   { color: #ffd700; text-shadow: 0 0 22px rgba(255,215,0,0.4); }
.stat-value.blue   { color: #70b8ff; }
.stat-value.green  { color: #80e8a8; }
.stat-value.orange { color: #ffaa44; }
.stat-value.red    { color: #ff7070; }
.stat-sublabel { font-size: 10px; color: #445070; margin-top: 3px; }
.accuracy-track { height: 4px; background: #1c2558; border-radius: 4px; margin-top: 10px; overflow: hidden; }
.accuracy-fill  { height: 100%; border-radius: 4px; background: linear-gradient(90deg, #1a5e30, #40c060); animation: barFill 1.5s ease-out both; }

/* ── Input labels ── */
.stSelectbox label, .stDateInput label { color: #7888b8 !important; font-size: 13px !important; }

/* ── TM superscript ── */
.tm-sup { font-size: 28%; vertical-align: super; letter-spacing: 2px; opacity: 0.75; font-family: 'Roboto Condensed', Arial, sans-serif; font-weight: 400; }

/* ── Golden football cursor ── */
html, body, * { cursor: url("data:image/svg+xml;charset=utf-8,%3Csvg xmlns='http://www.w3.org/2000/svg' width='22' height='22'%3E%3Cdefs%3E%3CradialGradient id='rg' cx='38%25' cy='32%25' r='60%25'%3E%3Cstop offset='0%25' stop-color='%23ffe566'/%3E%3Cstop offset='55%25' stop-color='%23d4af37'/%3E%3Cstop offset='100%25' stop-color='%23a07800'/%3E%3C/radialGradient%3E%3C/defs%3E%3Ccircle cx='11' cy='11' r='10' fill='url(%23rg)' stroke='%23886000' stroke-width='0.8'/%3E%3Cpath d='M11 4.5l1.8 3.5 3.5.5-2.6 2.3 1 3.7-3.7-2-3.7 2 1-3.7-2.6-2.3 3.5-.5z' fill='%23704000' opacity='0.5'/%3E%3C/svg%3E") 11 11, auto !important; }

/* ── Floating transparent footballs ── */
.fb {
    position: fixed; font-size: 56px; opacity: 0; pointer-events: none;
    z-index: 1; user-select: none; line-height: 1;
    animation: floatBall var(--fdur,22s) ease-in-out infinite var(--fdel,0s);
}
@keyframes floatBall {
    0%   { transform: translateY(102vh) rotate(0deg);   opacity: 0;     }
    6%   { opacity: 0.04; }
    50%  { transform: translateY(45vh)  rotate(270deg); opacity: 0.028; }
    94%  { opacity: 0.04; }
    100% { transform: translateY(-8vh)  rotate(540deg); opacity: 0;     }
}

/* ── Hide: top header bar, deploy button, sidebar, toolbar ── */
header[data-testid="stHeader"]         { display: none !important; }
[data-testid="stToolbar"]               { display: none !important; }
[data-testid="stDecoration"]            { display: none !important; }
.stDeployButton                         { display: none !important; }
section[data-testid="stSidebar"]        { display: none !important; }
button[data-testid="collapsedControl"]  { display: none !important; }
#MainMenu { display: none !important; }
footer    { display: none !important; }

/* ── Tournament stats + top scorers panel ── */
.ts-panel {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 16px;
    margin: 0 0 22px;
}
@media (max-width: 700px) { .ts-panel { grid-template-columns: 1fr; } }
.ts-card {
    background: linear-gradient(135deg, #08111f 0%, #0d1b30 100%);
    border: 1px solid #1a3050;
    border-radius: 12px;
    padding: 16px 18px 12px;
}
.ts-card-title {
    font-family: 'Bebas Neue', Impact, sans-serif;
    font-size: 15px;
    letter-spacing: 3px;
    color: #d4af37;
    margin-bottom: 10px;
    display: flex;
    align-items: center;
    gap: 7px;
}
.ts-row {
    display: flex;
    align-items: center;
    gap: 9px;
    padding: 5px 0;
    border-bottom: 1px solid rgba(255,255,255,0.045);
}
.ts-row:last-child { border-bottom: none; }
.ts-pos {
    font-family: 'Bebas Neue', Impact, sans-serif;
    font-size: 16px;
    width: 22px;
    color: #4a6a8a;
    flex-shrink: 0;
    text-align: center;
}
.ts-pos.gold   { color: #ffd700; }
.ts-pos.silver { color: #c0c0c0; }
.ts-pos.bronze { color: #cd7f32; }
.ts-flag-sm { flex-shrink: 0; }
.ts-flag-sm img { width: 24px; height: 16px; object-fit: cover; border-radius: 2px; border: 1px solid rgba(255,255,255,0.12); }
.ts-info { flex: 1; min-width: 0; }
.ts-name-sm {
    font-family: 'Roboto Condensed', Arial, sans-serif;
    font-weight: 700;
    font-size: 13px;
    color: #c8d4f0;
    white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
}
.ts-team-sm { font-size: 10px; color: #506080; font-family: 'Roboto Condensed',Arial,sans-serif; }
.ts-val {
    font-family: 'Bebas Neue', Impact, sans-serif;
    font-size: 20px;
    min-width: 28px;
    text-align: right;
    flex-shrink: 0;
}
.ts-val.goals   { color: #80e8a8; }
.ts-val.assists { color: #74c0fc; }
.ts-val.yellow  { color: #ffd43b; }
.ts-val.red     { color: #ff6b6b; }
.ts-val-lbl { font-size: 9px; color: #405060; letter-spacing: 1px; text-align: right; }

/* ── Player photo hover ── */
.ts-player-wrap {
    display: flex; align-items: center; gap: 7px; flex: 1; min-width: 0;
    position: relative;
}
.ts-photo-hover {
    position: absolute;
    bottom: 100%;
    left: 50%;
    transform: translateX(-50%);
    width: 72px;
    height: 72px;
    opacity: 0;
    pointer-events: none;
    transition: opacity 0.2s;
    z-index: 99;
    border-radius: 8px;
    overflow: hidden;
    box-shadow: 0 4px 20px rgba(0,0,0,0.7);
}
.ts-row:hover .ts-photo-hover { opacity: 1; }
.ts-photo-img {
    width: 100%; height: 100%; object-fit: cover;
}
</style>
""", unsafe_allow_html=True)

# ── FIFA Banner ─────────────────────────────────────────────────────────────
_logo_l = f'<img src="{FIFA_LOGO}" style="height:114px;filter:drop-shadow(0 0 24px rgba(212,175,55,0.8));">' if FIFA_LOGO else '<span style="font-size:70px;">⚽</span>'
_logo_r = f'<img src="{CONCEPT_LOGO}" style="height:132px;opacity:0.94;filter:drop-shadow(0 0 24px rgba(40,120,255,0.65));">' if CONCEPT_LOGO else ''
st.markdown(f"""
<div class="fifa-banner">
    <div class="banner-flex">
        <div class="banner-logo-left">{_logo_l}</div>
        <div class="banner-center">
            <div class="fifa-banner-logo">F I F A</div>
            <div class="fifa-title">WORLD CUP 2026<sup class="tm-sup">™</sup></div>
            <div class="fifa-hosts">
                <span class="canada">🍁 CANADA</span>
                <span class="sep">·</span>
                <span class="mexico">🌵 MEXICO</span>
                <span class="sep">·</span>
                <span class="usa">🦅 USA</span>
            </div>
            <div class="fifa-dates">June 11 &nbsp;—&nbsp; July 19, 2026 &nbsp;·&nbsp; Predictions Dashboard</div>
        </div>
        <div class="banner-logo-right">{_logo_r}</div>
    </div>
</div>
""", unsafe_allow_html=True)

# ── Floating footballs (transparent, decorative) ────────────────────────
st.markdown("""
<div class="fb" style="left:4%;  --fdur:22s; --fdel:0s">⚽</div>
<div class="fb" style="left:18%; --fdur:31s; --fdel:-12s">⚽</div>
<div class="fb" style="left:37%; --fdur:19s; --fdel:-5s">⚽</div>
<div class="fb" style="left:56%; --fdur:26s; --fdel:-18s">⚽</div>
<div class="fb" style="left:74%; --fdur:23s; --fdel:-8s">⚽</div>
<div class="fb" style="left:89%; --fdur:28s; --fdel:-15s">⚽</div>
""", unsafe_allow_html=True)

NZ_TZ = ZoneInfo("Pacific/Auckland")

def parse_time(val):
    try:
        if pd.isna(val): return None
        val = str(val).strip()
        if val in ["", "nan", "0"]: return None
        return pd.to_datetime(val).time()
    except:
        return None

def build_dt(row):
    if pd.isna(row["Date (NZDT)"]): return None
    base = row["Date (NZDT)"].date()
    t = row["ParsedTime"] or datetime.min.time()
    return datetime.combine(base, t).replace(tzinfo=NZ_TZ)

@st.cache_data(ttl=60)
def load_data(_mtime=0, _path=None):
    _target = _path or FILE_PATH
    _df = pd.read_excel(_target, sheet_name="Sheet1")
    _df["Date (NZDT)"] = pd.to_datetime(_df["Date (NZDT)"], errors="coerce")
    _df["ParsedTime"]  = _df["Time (NZDT)"].apply(parse_time)
    _df["DateTime"]    = _df.apply(build_dt, axis=1)
    return _df

# ── ESPN team-name normalisation map ─────────────────────────────────────────
_ESPN_MAP = {
    "united states":"United States","usa":"United States",
    "u.s.":"United States","u.s.a.":"United States",
    "türkiye":"Türkiye","turkey":"Türkiye",
    "korea republic":"South Korea","south korea":"South Korea",
    "republic of korea":"South Korea",
    "ir iran":"Iran","iran":"Iran",
    "côte d'ivoire":"Ivory Coast","ivory coast":"Ivory Coast",
    "bosnia & herzegovina":"Bosnia and Herzegovina",
    "bosnia and herzegovina":"Bosnia and Herzegovina",
    "north macedonia":"North Macedonia",
    "czechia":"Czech Republic","czech republic":"Czech Republic",
    "dr congo":"DR Congo","congo dr":"DR Congo",
    "saint kitts and nevis":"St Kitts and Nevis",
    "cabo verde":"Cape Verde",
    "cape verde":"Cape Verde",
}

def _norm_team(name):
    raw = " ".join(str(name).split()).strip()
    return _ESPN_MAP.get(raw.casefold(), raw)

def _team_key(name):
    return _norm_team(name).casefold()


def _is_draw_result(value):
    if value is None:
        return False
    try:
        if pd.isna(value):
            return False
    except Exception:
        pass

    s = " ".join(str(value).replace("—", "-").replace("–", "-").replace(":", "-").split()).casefold()
    if s in {"draw", "tie", "0-0", "0:0", "nil nil", "nil-nil"}:
        return True

    m = re.search(r'(?<!\d)(\d+)\s*[-]\s*(\d+)(?!\d)', s)
    return bool(m and m.group(1) == m.group(2))


def _status_text(*values):
    bits = []
    for v in values:
        if v is None:
            continue
        if isinstance(v, dict):
            for key in ("Description", "Name", "Text", "Status", "State", "Phase"):
                if v.get(key):
                    bits.append(str(v.get(key)))
            continue
        if isinstance(v, list):
            for item in v:
                bits.append(str(item))
            continue
        bits.append(str(v))
    return " ".join(bits).casefold()


def _is_liveish(status_text):
    return any(k in status_text for k in (
        "live", "in progress", "inprogress", "1st half", "2nd half",
        "half time", "halftime", "ht", "extra time", "penalty"
    ))


def _is_finishedish(status_text):
    return any(k in status_text for k in (
        "final", "finished", "completed", "full time", "ft", "aet", "post", "ended"
    ))


@st.cache_data(ttl=120)
def fetch_api_results():
    """Fetch only finished WC2026 results from official / fallback providers."""
    _out = {}
    _scores_out = {}
    _datetimes_out = {}  # team pair → kick-off datetime in NZT
    _now = datetime.now(NZ_TZ)
    _diso = _now.strftime("%Y-%m-%d")

    _urls = [
        "https://api.fifa.com/api/v3/calendar/matches?count=500&from=2026-06-11T00:00:00Z&to=2026-07-20T23:59:59Z&language=en",
        "https://api.fifa.com/api/v3/calendar/matches?count=500&from=2026-06-11T00:00:00Z&to=2026-07-20T23:59:59Z&language=en&sort=Date",
    ]

    def _first_desc(val):
        if isinstance(val, list) and val:
            v = val[0]
            if isinstance(v, dict):
                return v.get("Description") or v.get("Name") or v.get("Text") or ""
            return str(v)
        if isinstance(val, dict):
            return val.get("Description") or val.get("Name") or val.get("Text") or ""
        if val is None:
            return ""
        return str(val)

    def _team_name(block):
        if not isinstance(block, dict):
            return ""
        for k in ("ShortClubName", "TeamName", "Name", "Abbreviation"):
            v = block.get(k)
            if v:
                desc = _first_desc(v)
                if desc:
                    return desc.strip()
        return ""

    def _score(block, primary_key, fallback_key):
        try:
            if isinstance(block, dict):
                v = block.get(primary_key)
                if v is None:
                    v = block.get(fallback_key)
                if v is None and "Score" in block:
                    v = block.get("Score")
                if v is None:
                    return None
                return int(v)
        except:
            pass
        return None

    def _consume(payload):
        for _ev in payload.get("Results") or []:
            _h = _team_name(_ev.get("Home"))
            _a = _team_name(_ev.get("Away"))
            if not _h or not _a:
                continue

            _status = _status_text(
                _ev.get("Status"),
                _ev.get("MatchStatus"),
                _ev.get("status"),
                _ev.get("Phase"),
                _ev.get("Progress"),
                _ev.get("State"),
                _ev.get("GameStatus"),
                _ev.get("MatchState"),
                _ev.get("CompetitionStatus"),
                _ev.get("StatusDescription"),
                _ev.get("Description"),
            )

            # Skip anything that is not explicitly finished.
            if _is_liveish(_status) or not _is_finishedish(_status):
                continue

            hs = _score(_ev.get("Home"), "Score", "HomeTeamScore")
            as_ = _score(_ev.get("Away"), "Score", "AwayTeamScore")
            if hs is None:
                hs = _score(_ev, "HomeTeamScore", "HomeScore")
            if as_ is None:
                as_ = _score(_ev, "AwayTeamScore", "AwayScore")
            if hs is None or as_ is None:
                continue

            _n0 = _norm_team(_h)
            _n1 = _norm_team(_a)
            _w = _n0 if hs > as_ else _n1 if as_ > hs else "Draw"
            _out[(_team_key(_n0), _team_key(_n1))] = _w
            _out[(_team_key(_n1), _team_key(_n0))] = _w
            # Store scores so match cards can display "2 - 1" etc.
            _scores_out[(_team_key(_n0), _team_key(_n1))] = (hs, as_)
            _scores_out[(_team_key(_n1), _team_key(_n0))] = (as_, hs)
            # Store API kick-off datetime (authoritative — overrides Excel dates)
            for _dk in ("Date", "MatchDateTime", "LocalDate", "UtcDate", "DateUtc", "KickOffTime"):
                _dv = _ev.get(_dk)
                if _dv:
                    try:
                        _dt_parsed = datetime.fromisoformat(str(_dv).replace("Z", "+00:00"))
                        _dt_nzt = _dt_parsed.astimezone(NZ_TZ)
                        _datetimes_out[(_team_key(_n0), _team_key(_n1))] = _dt_nzt
                        _datetimes_out[(_team_key(_n1), _team_key(_n0))] = _dt_nzt
                        break
                    except Exception:
                        pass

    for _url in _urls:
        try:
            _resp = req.get(_url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
            if _resp.status_code != 200:
                continue
            _consume(_resp.json())
            break  # got a valid response from FIFA, no need to try the second URL
        except:
            continue

    # Always run ESPN for a full date range — FIFA may have missed matches,
    # and ESPN covers the entire tournament history.
    # Generate all dates from tournament start through today + 1
    _tourney_start = datetime(2026, 6, 11, tzinfo=NZ_TZ).date()
    _scan_end      = (_now + timedelta(days=1)).date()
    _all_days = []
    _d = _tourney_start
    while _d <= _scan_end:
        _all_days.append(_d.strftime("%Y%m%d"))
        _d += timedelta(days=1)

    _espn_slugs = [
        "fifa.worldcup",
        "fifa.world",
        "global.2026-fifa-world-cup",
        "fifa.worldcup.2026",
    ]
    # Find the working slug first using today's date
    _working_slug = None
    for _slug in _espn_slugs:
        try:
            _test = req.get(
                f"https://site.api.espn.com/apis/site/v2/sports/soccer/{_slug}/scoreboard?dates={_now.strftime('%Y%m%d')}",
                timeout=8, headers={"User-Agent": "Mozilla/5.0"}
            )
            if _test.status_code == 200 and _test.json().get("events") is not None:
                _working_slug = _slug
                break
        except:
            continue

    if _working_slug:
        for _day in _all_days:
            try:
                _url  = f"https://site.api.espn.com/apis/site/v2/sports/soccer/{_working_slug}/scoreboard?dates={_day}"
                _resp = req.get(_url, timeout=8, headers={"User-Agent":"Mozilla/5.0"})
                if _resp.status_code != 200:
                    continue
                _ev_list = _resp.json().get("events", [])
                for _ev in _ev_list:
                    _comp   = (_ev.get("competitions") or [{}])[0]
                    _status = _comp.get("status", {}).get("type", {})
                    _done   = bool(_status.get("completed", False) or _status.get("state") == "post")
                    if not _done:
                        continue

                    _teams = _comp.get("competitors", [])
                    if len(_teams) < 2:
                        continue

                    _names = [t.get("team", {}).get("displayName", "") for t in _teams]
                    _sc = []
                    for _t in _teams:
                        try:
                            _sc.append(int(_t.get("score", 0) or 0))
                        except:
                            _sc.append(0)

                    _n0 = _norm_team(_names[0])
                    _n1 = _norm_team(_names[1])
                    _k0, _k1 = _team_key(_n0), _team_key(_n1)
                    # Fill result + score gaps (don't overwrite FIFA data)
                    if (_k0, _k1) not in _out:
                        _w = _n0 if _sc[0] > _sc[1] else _n1 if _sc[1] > _sc[0] else "Draw"
                        _out[(_k0, _k1)] = _w
                        _out[(_k1, _k0)] = _w
                    # Always fill score gaps from ESPN
                    if (_k0, _k1) not in _scores_out:
                        _scores_out[(_k0, _k1)] = (_sc[0], _sc[1])
                        _scores_out[(_k1, _k0)] = (_sc[1], _sc[0])
                    # Fill datetime gaps
                    if (_k0, _k1) not in _datetimes_out:
                        _raw_dt = _ev.get("date") or _comp.get("date")
                        if _raw_dt:
                            try:
                                _dt_p = datetime.fromisoformat(str(_raw_dt).replace("Z","+00:00"))
                                _dt_nzt = _dt_p.astimezone(NZ_TZ)
                                _datetimes_out[(_k0, _k1)] = _dt_nzt
                                _datetimes_out[(_k1, _k0)] = _dt_nzt
                            except Exception:
                                pass
            except:
                continue

    # TheSportsDB — fills any remaining gaps across all past tournament dates
    _fin_stati = {"match finished","ft","aet","pen","ap"}
    _past_days = []
    _d2 = _tourney_start
    while _d2 <= _now.date():
        _past_days.append(_d2.strftime("%Y-%m-%d"))
        _d2 += timedelta(days=1)

    for _tsdb_date in _past_days:
        # Skip if we already have all matches for this date from ESPN/FIFA
        try:
            _url  = f"https://www.thesportsdb.com/api/v1/json/3/eventsday.php?d={_tsdb_date}&l=FIFA+World+Cup"
            _resp = req.get(_url, timeout=8, headers={"User-Agent":"Mozilla/5.0"})
            if _resp.status_code != 200:
                continue
            for _ev in (_resp.json().get("events") or []):
                _st = (_ev.get("strStatus") or "").lower().strip()
                _hs_raw = _ev.get("intHomeScore")
                _as_raw = _ev.get("intAwayScore")
                if _st not in _fin_stati or _hs_raw is None:
                    continue
                _h = _ESPN_MAP.get((_ev.get("strHomeTeam") or "").lower(), _ev.get("strHomeTeam", ""))
                _a = _ESPN_MAP.get((_ev.get("strAwayTeam") or "").lower(), _ev.get("strAwayTeam", ""))
                _k0, _k1 = _h.lower(), _a.lower()
                try: _hs, _as = int(_hs_raw), int(_as_raw)
                except: _hs, _as = 0, 0
                if (_k0, _k1) not in _out:
                    _w = _h if _hs>_as else _a if _as>_hs else "Draw"
                    _out[(_k0, _k1)] = _w
                    _out[(_k1, _k0)] = _w
                if (_k0, _k1) not in _scores_out:
                    _scores_out[(_k0, _k1)] = (_hs, _as)
                    _scores_out[(_k1, _k0)] = (_as, _hs)
                if (_k0, _k1) not in _datetimes_out:
                    _ts = _ev.get("strTimestamp") or _ev.get("strDateTimeUTC")
                    if _ts:
                        try:
                            _dt_p = datetime.fromisoformat(str(_ts).replace("Z","+00:00"))
                            _dt_nzt = _dt_p.astimezone(NZ_TZ)
                            _datetimes_out[(_k0, _k1)] = _dt_nzt
                            _datetimes_out[(_k1, _k0)] = _dt_nzt
                        except Exception:
                            pass
        except:
            continue

    # Scoreaxis — last resort for any still-missing matches
    try:
        _url  = f"https://www.scoreaxis.com/api/live-events?sport=1&date={_diso}&timeZone=0"
        _resp = req.get(_url, timeout=6, headers={"User-Agent":"Mozilla/5.0"})
        if _resp.status_code == 200:
            for _ev in (_resp.json().get("data",{}).get("events") or []):
                if not _ev.get("statusFinished"): continue
                _h = _ESPN_MAP.get((_ev.get("homeTeamName") or "").lower(), _ev.get("homeTeamName",""))
                _a = _ESPN_MAP.get((_ev.get("awayTeamName") or "").lower(), _ev.get("awayTeamName",""))
                _k0, _k1 = _h.lower(), _a.lower()
                if (_k0, _k1) not in _out:
                    try: _hs = int(_ev.get("homeScore",0)); _as = int(_ev.get("awayScore",0))
                    except: _hs, _as = 0, 0
                    _w = _h if _hs>_as else _a if _as>_hs else "Draw"
                    _out[(_k0, _k1)] = _w
                    _out[(_k1, _k0)] = _w
                    _scores_out[(_k0, _k1)] = (_hs, _as)
                    _scores_out[(_k1, _k0)] = (_as, _hs)
    except:
        pass

    return _out, _scores_out, _datetimes_out


@st.cache_data(ttl=3600)   # schedule rarely changes — cache 1 hour
def fetch_schedule():
    """
    Fetch kick-off datetimes (in NZT) for ALL WC2026 matches — past, live, upcoming.
    Returns dict: (team_key_a, team_key_b) → datetime (NZT, timezone-aware)
    Sources: FIFA full calendar → ESPN date range → TheSportsDB.
    Excel dates/times are NEVER used for display; this is the authoritative source.
    """
    _out = {}
    _now = datetime.now(NZ_TZ)

    def _store_dt(name_a, name_b, raw_dt_str):
        if not raw_dt_str:
            return
        try:
            _dt = datetime.fromisoformat(str(raw_dt_str).replace("Z", "+00:00"))
            _nzt = _dt.astimezone(NZ_TZ)
            _k0, _k1 = _team_key(_norm_team(name_a)), _team_key(_norm_team(name_b))
            if _k0 and _k1:
                _out[(_k0, _k1)] = _nzt
                _out[(_k1, _k0)] = _nzt
        except Exception:
            pass

    # ── Source 1: FIFA full calendar (all statuses, not just finished) ──────
    _fifa_urls = [
        "https://api.fifa.com/api/v3/calendar/matches?count=500&from=2026-06-11T00:00:00Z&to=2026-07-20T23:59:59Z&language=en",
        "https://api.fifa.com/api/v3/calendar/matches?count=500&from=2026-06-11T00:00:00Z&to=2026-07-20T23:59:59Z&language=en&sort=Date",
    ]
    def _first_desc(val):
        if isinstance(val, list) and val:
            v = val[0]
            return (v.get("Description") or v.get("Name") or v.get("Text") or "") if isinstance(v, dict) else str(v)
        if isinstance(val, dict):
            return val.get("Description") or val.get("Name") or val.get("Text") or ""
        return str(val) if val else ""

    def _tname(block):
        if not isinstance(block, dict): return ""
        for k in ("ShortClubName", "TeamName", "Name", "Abbreviation"):
            v = block.get(k)
            if v:
                d = _first_desc(v)
                if d: return d.strip()
        return ""

    for _url in _fifa_urls:
        try:
            _r = req.get(_url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
            if _r.status_code != 200:
                continue
            for _ev in (_r.json().get("Results") or []):
                _h, _a = _tname(_ev.get("Home")), _tname(_ev.get("Away"))
                if not _h or not _a:
                    continue
                for _dk in ("Date", "MatchDateTime", "LocalDate", "UtcDate", "DateUtc", "KickOffTime"):
                    if _ev.get(_dk):
                        _store_dt(_h, _a, _ev[_dk])
                        break
            if _out:
                break
        except Exception:
            continue

    # ── Source 2: ESPN — scan full tournament date range ────────────────────
    _tourney_start = datetime(2026, 6, 11, tzinfo=NZ_TZ).date()
    _tourney_end   = datetime(2026, 7, 20, tzinfo=NZ_TZ).date()
    _espn_slugs = ["fifa.worldcup", "fifa.world", "global.2026-fifa-world-cup", "fifa.worldcup.2026"]

    # Find working slug
    _working_slug = None
    for _slug in _espn_slugs:
        try:
            _t = req.get(
                f"https://site.api.espn.com/apis/site/v2/sports/soccer/{_slug}/scoreboard?dates={_now.strftime('%Y%m%d')}",
                timeout=6, headers={"User-Agent": "Mozilla/5.0"}
            )
            if _t.status_code == 200 and _t.json().get("events") is not None:
                _working_slug = _slug
                break
        except Exception:
            continue

    if _working_slug:
        _d = _tourney_start
        while _d <= _tourney_end:
            try:
                _r = req.get(
                    f"https://site.api.espn.com/apis/site/v2/sports/soccer/{_working_slug}/scoreboard?dates={_d.strftime('%Y%m%d')}",
                    timeout=8, headers={"User-Agent": "Mozilla/5.0"}
                )
                if _r.status_code == 200:
                    for _ev in (_r.json().get("events") or []):
                        _comp = (_ev.get("competitions") or [{}])[0]
                        _teams = _comp.get("competitors", [])
                        if len(_teams) < 2:
                            continue
                        _names = [t.get("team", {}).get("displayName", "") for t in _teams]
                        _raw_dt = _ev.get("date") or _comp.get("date")
                        if (_team_key(_norm_team(_names[0])), _team_key(_norm_team(_names[1]))) not in _out:
                            _store_dt(_names[0], _names[1], _raw_dt)
            except Exception:
                pass
            _d += timedelta(days=1)

    # ── Source 3: TheSportsDB — fill any remaining gaps ─────────────────────
    _d = _tourney_start
    while _d <= _now.date():
        try:
            _r = req.get(
                f"https://www.thesportsdb.com/api/v1/json/3/eventsday.php?d={_d.strftime('%Y-%m-%d')}&l=FIFA+World+Cup",
                timeout=6, headers={"User-Agent": "Mozilla/5.0"}
            )
            if _r.status_code == 200:
                for _ev in (_r.json().get("events") or []):
                    _h = _ESPN_MAP.get((_ev.get("strHomeTeam") or "").lower(), _ev.get("strHomeTeam", ""))
                    _a = _ESPN_MAP.get((_ev.get("strAwayTeam") or "").lower(), _ev.get("strAwayTeam", ""))
                    _k0, _k1 = _team_key(_h), _team_key(_a)
                    if (_k0, _k1) not in _out:
                        _store_dt(_h, _a, _ev.get("strTimestamp") or _ev.get("strDateTimeUTC"))
        except Exception:
            pass
        _d += timedelta(days=1)

    return _out


@st.cache_data(ttl=120)
def fetch_match_goalscorers():
    """
    Returns {(k1,k2): {"home":[{name,minute,og}], "away":[...]}}
    Tries FIFA MatchEvents → ESPN competition details.
    """
    _out = {}
    _now = datetime.now(NZ_TZ)

    def _d(val):
        if isinstance(val, list) and val:
            v = val[0]
            return (v.get("Description","") or v.get("Name","") if isinstance(v, dict) else str(v))
        if isinstance(val, dict): return val.get("Description","") or val.get("Name","")
        return str(val) if val else ""

    def _tname(b):
        if not isinstance(b, dict): return ""
        for k in ("ShortClubName","TeamName","Name"):
            v = b.get(k)
            if v:
                t = _d(v)
                if t: return t.strip()
        return ""

    # ── FIFA ────────────────────────────────────────────────────────────────
    try:
        _r = req.get(
            "https://api.fifa.com/api/v3/calendar/matches?count=500"
            "&from=2026-06-11T00:00:00Z&to=2026-07-20T23:59:59Z&language=en",
            timeout=12, headers={"User-Agent": "Mozilla/5.0"}
        )
        if _r.status_code == 200:
            for _ev in (_r.json().get("Results") or []):
                _h_blk = _ev.get("Home") or {}
                _a_blk = _ev.get("Away") or {}
                _h = _norm_team(_tname(_h_blk))
                _a = _norm_team(_tname(_a_blk))
                if not _h or not _a: continue
                _k0, _k1 = _team_key(_h), _team_key(_a)
                _home_id = str(_h_blk.get("IdTeam") or _h_blk.get("TeamId") or "")
                _away_id = str(_a_blk.get("IdTeam") or _a_blk.get("TeamId") or "")
                # Store by team key — safe even when TeamId is missing
                _by_team = {_k0: [], _k1: []}
                for _me in (_ev.get("MatchEvents") or []):
                    _tid  = str(_me.get("IdTeam") or _me.get("TeamId") or "")
                    _type = str(_me.get("TypeId") or _me.get("Type") or "")
                    _txt  = str(_me.get("EventDescription") or _me.get("TypeName") or "").lower()
                    if _type not in ("0","1","2") and "goal" not in _txt: continue
                    _og   = _type == "1" or "own" in _txt
                    _sraw = _me.get("Scorer") or _me.get("PlayerName") or {}
                    _sname = _d(_sraw.get("ShortName") or _sraw.get("Name") or "") if isinstance(_sraw, dict) else str(_sraw)
                    _min   = str(_me.get("MatchMinute") or _me.get("Minute") or "?")
                    _entry = {"name": _sname.strip(), "minute": _min, "og": _og}
                    if _tid == _home_id and _home_id:
                        _by_team[_k0].append(_entry)
                    elif _tid == _away_id and _away_id:
                        _by_team[_k1].append(_entry)
                    else:
                        # TeamId unknown — skip (will be filled by ESPN fallback)
                        pass
                if any(_by_team.values()):
                    _out[(_k0,_k1)] = _by_team
                    _out[(_k1,_k0)] = _by_team
    except Exception:
        pass

    # ── ESPN competition details fallback ────────────────────────────────────
    try:
        _espn_slugs = ["fifa.worldcup","fifa.world","global.2026-fifa-world-cup","fifa.worldcup.2026"]
        _slug = None
        for _s in _espn_slugs:
            _t = req.get(f"https://site.api.espn.com/apis/site/v2/sports/soccer/{_s}/scoreboard?dates={_now.strftime('%Y%m%d')}",
                         timeout=6, headers={"User-Agent":"Mozilla/5.0"})
            if _t.status_code == 200 and _t.json().get("events") is not None:
                _slug = _s; break
        if _slug:
            _d2 = datetime(2026, 6, 11).date()
            while _d2 <= _now.date():
                _r2 = req.get(f"https://site.api.espn.com/apis/site/v2/sports/soccer/{_slug}/scoreboard?dates={_d2.strftime('%Y%m%d')}",
                              timeout=8, headers={"User-Agent":"Mozilla/5.0"})
                if _r2.status_code == 200:
                    for _ev in (_r2.json().get("events") or []):
                        _comp  = (_ev.get("competitions") or [{}])[0]
                        _teams = _comp.get("competitors",[])
                        if len(_teams) < 2: continue

                        _k0 = _team_key(_norm_team(_teams[0].get("team",{}).get("displayName","")))
                        _k1 = _team_key(_norm_team(_teams[1].get("team",{}).get("displayName","")))
                        if (_k0,_k1) in _out: continue

                        # Build homeAway string → team_key map (e.g. "home"→"brazil", "away"→"japan")
                        _ha_to_key = {}
                        _score_map = {}
                        for _tc in _teams:
                            _tkey = _team_key(_norm_team(_tc.get("team",{}).get("displayName","")))
                            _ha_val = _tc.get("homeAway","").lower()
                            if _ha_val:
                                _ha_to_key[_ha_val] = _tkey
                            try: _score_map[_tkey] = int(_tc.get("score",0) or 0)
                            except: _score_map[_tkey] = 0

                        _status3 = _comp.get("status",{}).get("type",{})
                        _done3 = bool(_status3.get("completed") or _status3.get("state")=="post")
                        if not _done3: continue

                        _by_team = {_k0: [], _k1: []}
                        for _sc in (_comp.get("details") or []):
                            _txt2 = str(_sc.get("type",{}).get("text","")).lower()
                            if not any(g in _txt2 for g in ("goal","penalty")): continue
                            _inv  = _sc.get("athletesInvolved") or []
                            _sn   = _inv[0].get("displayName","") if _inv else ""
                            _min2 = _sc.get("clock",{}).get("displayValue","?")
                            _og2  = "own" in _txt2
                            # detail.homeAway = the side that SCORED this goal (benefiting team)
                            _det_ha = _sc.get("homeAway","").lower()
                            _scorer_key = _ha_to_key.get(_det_ha)
                            if _scorer_key is None:
                                # fallback: use competitor index
                                _scorer_key = _k0
                            if _scorer_key in _by_team:
                                _by_team[_scorer_key].append({"name":_sn,"minute":_min2,"og":_og2})

                        # Only gap-fill if we have SOME named scorers (don't fake missing teams)
                        _total_found = sum(len(v) for v in _by_team.values())
                        if _total_found > 0:
                            for _tkey, _goals in _by_team.items():
                                _expected = _score_map.get(_tkey, 0)
                                if 0 < len(_goals) < _expected:
                                    for _ in range(_expected - len(_goals)):
                                        _goals.append({"name":"—","minute":"?","og":False})

                        if any(_by_team.values()):
                            # Store by team key directly — avoids positional confusion
                            _out[(_k0,_k1)] = _by_team   # {team_key: [goals]}
                            _out[(_k1,_k0)] = _by_team   # same dict, both orderings
                _d2 += timedelta(days=1)
    except Exception:
        pass

    return _out


@st.cache_data(ttl=300)
def fetch_player_stats():
    """
    Returns dict with keys: top_scorers, top_assists, yellow_cards, red_cards
    Each is a list of {name, team, count, flag_code}
    """
    _now = datetime.now(NZ_TZ)
    _result = {"top_scorers":[], "top_assists":[], "yellow_cards":[], "red_cards":[]}

    def _desc(val):
        if isinstance(val, list) and val:
            v = val[0]; return (v.get("Description","") if isinstance(v,dict) else str(v))
        if isinstance(val, dict): return val.get("Description","") or val.get("Name","")
        return str(val) if val else ""

    # ── Build from FIFA match events ─────────────────────────────────────────
    _scorers   = {}   # player → {team, goals}
    _assisters = {}
    _yellows   = {}
    _reds      = {}

    try:
        _r = req.get(
            "https://api.fifa.com/api/v3/calendar/matches?count=500"
            "&from=2026-06-11T00:00:00Z&to=2026-07-20T23:59:59Z&language=en",
            timeout=12, headers={"User-Agent":"Mozilla/5.0"}
        )
        if _r.status_code == 200:
            for _ev in (_r.json().get("Results") or []):
                def _tname2(b):
                    if not isinstance(b, dict): return ""
                    for k in ("ShortClubName","TeamName","Name"):
                        v = b.get(k)
                        if v:
                            t = _desc(v)
                            if t: return t.strip()
                    return ""
                _h_blk = _ev.get("Home") or {}
                _a_blk = _ev.get("Away") or {}
                _home_id = str(_h_blk.get("IdTeam") or _h_blk.get("TeamId") or "")
                _h_name  = _norm_team(_tname2(_h_blk))
                _a_name  = _norm_team(_tname2(_a_blk))

                for _me in (_ev.get("MatchEvents") or []):
                    _tid   = str(_me.get("IdTeam") or _me.get("TeamId") or "")
                    _team  = _h_name if _tid == _home_id else _a_name
                    _type  = str(_me.get("TypeId") or "")
                    _txt   = str(_me.get("EventDescription") or _me.get("TypeName") or "").lower()

                    def _pname(key):
                        raw = _me.get(key) or {}
                        if isinstance(raw, dict):
                            return _desc(raw.get("ShortName") or raw.get("Name") or "")
                        return str(raw)

                    # Goals (0=goal, 2=penalty)
                    if _type in ("0","2") or ("goal" in _txt and "own" not in _txt):
                        _sn = _pname("Scorer") or _pname("PlayerName")
                        if _sn:
                            if _sn not in _scorers: _scorers[_sn] = {"team":_team,"count":0}
                            _scorers[_sn]["count"] += 1

                    # Assists
                    if _type == "0" or "goal" in _txt:
                        _an = _pname("Assist") or _pname("AssistPlayer")
                        if _an:
                            if _an not in _assisters: _assisters[_an] = {"team":_team,"count":0}
                            _assisters[_an]["count"] += 1

                    # Yellow cards (TypeId 4 or 5)
                    if _type in ("4","5") or "yellow" in _txt:
                        _pn = _pname("Player") or _pname("PlayerName")
                        if _pn:
                            if _pn not in _yellows: _yellows[_pn] = {"team":_team,"count":0}
                            _yellows[_pn]["count"] += 1

                    # Red cards (TypeId 6 or 7)
                    if _type in ("6","7") or "red" in _txt:
                        _pn = _pname("Player") or _pname("PlayerName")
                        if _pn:
                            if _pn not in _reds: _reds[_pn] = {"team":_team,"count":0}
                            _reds[_pn]["count"] += 1
    except Exception:
        pass

    def _to_list(d, top=5):
        rows = [{"name":n,"team":v["team"],"count":v["count"],
                 "flag_code":TEAM_FLAG_MAP.get(v["team"],""),
                 "photo": v.get("photo","")} for n,v in d.items()]
        rows.sort(key=lambda x: -x["count"])
        return rows[:top]

    _result["top_scorers"]   = _to_list(_scorers)
    _result["top_assists"]   = _to_list(_assisters)
    _result["yellow_cards"]  = _to_list(_yellows)
    _result["red_cards"]     = _to_list(_reds)

    # ── Build stats from goalscorer data we already fetched (avoids duplicate API calls) ──
    # fetch_player_stats is called after fetch_match_goalscorers, so we reuse that data
    # by re-fetching the same ESPN scoreboard endpoint (now with working slug detection)
    _espn_slugs = ["fifa.worldcup","fifa.world","global.2026-fifa-world-cup","fifa.worldcup.2026"]
    _working = None
    for _s in _espn_slugs:
        for _test_date in [_now.strftime('%Y%m%d'), "20260628", "20260625", "20260617"]:
            try:
                _t = req.get(
                    f"https://site.api.espn.com/apis/site/v2/sports/soccer/{_s}/scoreboard?dates={_test_date}",
                    timeout=6, headers={"User-Agent":"Mozilla/5.0"}
                )
                if _t.status_code == 200 and _t.json().get("events") is not None:
                    _working = _s; break
            except Exception:
                continue
        if _working: break

    if _working:
        _espn_scorers, _espn_assists, _espn_yellows, _espn_reds = {}, {}, {}, {}
        _scan_d = datetime(2026, 6, 11).date()

        def _add(d, name, team, fc, photo):
            if name not in d:
                d[name] = {"team": team, "count": 0, "flag_code": fc, "photo": photo}
            d[name]["count"] += 1

        while _scan_d <= _now.date():
            try:
                _rr = req.get(
                    f"https://site.api.espn.com/apis/site/v2/sports/soccer/{_working}/scoreboard?dates={_scan_d.strftime('%Y%m%d')}",
                    timeout=8, headers={"User-Agent":"Mozilla/5.0"}
                )
                if _rr.status_code == 200:
                    for _ev in (_rr.json().get("events") or []):
                        _comp = (_ev.get("competitions") or [{}])[0]
                        if not (_comp.get("status",{}).get("type",{}).get("completed") or
                                _comp.get("status",{}).get("type",{}).get("state") == "post"):
                            continue
                        _ha_to_team = {}
                        for _tc in (_comp.get("competitors") or []):
                            _ha_ = _tc.get("homeAway","")
                            _nt_ = _norm_team(_tc.get("team",{}).get("displayName",""))
                            _ha_to_team[_ha_] = _nt_
                        for _det in (_comp.get("details") or []):
                            _txt_ = str(_det.get("type",{}).get("text","")).lower()
                            _inv_ = _det.get("athletesInvolved") or []
                            _ha_  = _det.get("homeAway","")
                            _nt_  = _ha_to_team.get(_ha_,"")
                            _fc_  = TEAM_FLAG_MAP.get(_nt_,"")
                            def _ph_(a):
                                if not a: return ""
                                _hs_ = a.get("headshot")
                                if isinstance(_hs_, dict): return _hs_.get("href","")
                                _id_ = a.get("id","")
                                return f"https://a.espncdn.com/combiner/i?img=/i/headshots/soccer/players/full/{_id_}.png&w=96&h=70" if _id_ else ""
                            if ("goal" in _txt_ or "penalty" in _txt_) and _inv_:
                                _pn_ = _inv_[0].get("displayName","")
                                if _pn_: _add(_espn_scorers, _pn_, _nt_, _fc_, _ph_(_inv_[0]))
                                if len(_inv_) > 1 and "own" not in _txt_:
                                    _an_ = _inv_[1].get("displayName","")
                                    if _an_: _add(_espn_assists, _an_, _nt_, _fc_, _ph_(_inv_[1]))
                            elif "yellow card" in _txt_ and _inv_:
                                _pn_ = _inv_[0].get("displayName","")
                                if _pn_: _add(_espn_yellows, _pn_, _nt_, _fc_, _ph_(_inv_[0]))
                            elif "red card" in _txt_ and _inv_:
                                _pn_ = _inv_[0].get("displayName","")
                                if _pn_: _add(_espn_reds, _pn_, _nt_, _fc_, _ph_(_inv_[0]))
            except Exception:
                pass
            _scan_d += timedelta(days=1)

        if _espn_scorers: _result["top_scorers"]  = _to_list(_espn_scorers)
        if _espn_assists: _result["top_assists"]   = _to_list(_espn_assists)
        if _espn_yellows: _result["yellow_cards"]  = _to_list(_espn_yellows)
        if _espn_reds:    _result["red_cards"]     = _to_list(_espn_reds)

    return _result

    if _working:
        from concurrent.futures import ThreadPoolExecutor, as_completed

        _espn_scorers, _espn_assists, _espn_yellows, _espn_reds = {}, {}, {}, {}
        _all_days = []
        _scan_d = datetime(2026, 6, 11).date()
        while _scan_d <= _now.date():
            _all_days.append(_scan_d.strftime('%Y%m%d'))
            _scan_d += timedelta(days=1)

        def _fetch_day_stats(day_str):
            try:
                _r = req.get(
                    f"https://site.api.espn.com/apis/site/v2/sports/soccer/{_working}/scoreboard?dates={day_str}",
                    timeout=8, headers={"User-Agent":"Mozilla/5.0"}
                )
                if _r.status_code != 200: return []
                results = []
                for _ev in (_r.json().get("events") or []):
                    _comp = (_ev.get("competitions") or [{}])[0]
                    _team_by_ha = {}
                    for _tc in (_comp.get("competitors") or []):
                        _ha_ = _tc.get("homeAway","")
                        _nt_ = _norm_team(_tc.get("team",{}).get("displayName",""))
                        _team_by_ha[_ha_] = _nt_
                    for _det in (_comp.get("details") or []):
                        _txt_ = str(_det.get("type",{}).get("text","")).lower()
                        _inv_ = _det.get("athletesInvolved") or []
                        _ha_  = _det.get("homeAway","")
                        _nt_  = _team_by_ha.get(_ha_,"")
                        _fc_  = TEAM_FLAG_MAP.get(_nt_,"")
                        def _ph_(a):
                            if not a: return ""
                            _hs_ = a.get("headshot")
                            if isinstance(_hs_, dict): return _hs_.get("href","")
                            _id_ = a.get("id","")
                            return f"https://a.espncdn.com/combiner/i?img=/i/headshots/soccer/players/full/{_id_}.png&w=96&h=70" if _id_ else ""
                        if "goal" in _txt_ or "penalty" in _txt_:
                            if _inv_:
                                _pn_ = _inv_[0].get("displayName","")
                                if _pn_: results.append(("goal", _pn_, _nt_, _fc_, _ph_(_inv_[0])))
                                if len(_inv_) > 1 and "own" not in _txt_:
                                    _an_ = _inv_[1].get("displayName","")
                                    if _an_: results.append(("assist", _an_, _nt_, _fc_, _ph_(_inv_[1])))
                        if "yellow card" in _txt_ and _inv_:
                            _pn_ = _inv_[0].get("displayName","")
                            if _pn_: results.append(("yellow", _pn_, _nt_, _fc_, _ph_(_inv_[0])))
                        if "red card" in _txt_ and _inv_:
                            _pn_ = _inv_[0].get("displayName","")
                            if _pn_: results.append(("red", _pn_, _nt_, _fc_, _ph_(_inv_[0])))
                return results
            except Exception:
                return []

        with ThreadPoolExecutor(max_workers=6) as _pool:
            _futures = {_pool.submit(_fetch_day_stats, d): d for d in _all_days}
            for _fut in as_completed(_futures):
                for _evt_type, _pname_, _nt_, _fc_, _photo_ in (_fut.result() or []):
                    if _evt_type == "goal":
                        if _pname_ not in _espn_scorers:
                            _espn_scorers[_pname_] = {"team":_nt_,"count":0,"flag_code":_fc_,"photo":_photo_}
                        _espn_scorers[_pname_]["count"] += 1
                    elif _evt_type == "assist":
                        if _pname_ not in _espn_assists:
                            _espn_assists[_pname_] = {"team":_nt_,"count":0,"flag_code":_fc_,"photo":_photo_}
                        _espn_assists[_pname_]["count"] += 1
                    elif _evt_type == "yellow":
                        if _pname_ not in _espn_yellows:
                            _espn_yellows[_pname_] = {"team":_nt_,"count":0,"flag_code":_fc_,"photo":_photo_}
                        _espn_yellows[_pname_]["count"] += 1
                    elif _evt_type == "red":
                        if _pname_ not in _espn_reds:
                            _espn_reds[_pname_] = {"team":_nt_,"count":0,"flag_code":_fc_,"photo":_photo_}
                        _espn_reds[_pname_]["count"] += 1

        if _espn_scorers: _result["top_scorers"]  = _to_list(_espn_scorers)
        if _espn_assists: _result["top_assists"]   = _to_list(_espn_assists)
        if _espn_yellows: _result["yellow_cards"]  = _to_list(_espn_yellows)
        if _espn_reds:    _result["red_cards"]     = _to_list(_espn_reds)

    return _result


@st.cache_data(ttl=15)
def fetch_live_scores():
    """Fetch live scores for ongoing matches.
    Tries FIFA first, then ESPN, then Scoreaxis fallback.
    Returns mapping of (team_key_a, team_key_b) -> {"home": int, "away": int, "status": str}
    """
    _out = {}
    _now = datetime.now(NZ_TZ)
    _days = [(_now.date() + timedelta(days=d)).strftime("%Y%m%d") for d in (-1, 0, 1)]
    _iso_days = [(_now.date() + timedelta(days=d)).strftime("%Y-%m-%d") for d in (-1, 0, 1)]

    _fifa_urls = [
        "https://api.fifa.com/api/v3/calendar/matches?count=500&from=2026-06-11T00:00:00Z&to=2026-07-20T23:59:59Z&language=en",
        "https://api.fifa.com/api/v3/calendar/matches?count=500&from=2026-06-11T00:00:00Z&to=2026-07-20T23:59:59Z&language=en&sort=Date",
    ]

    def _first_desc(val):
        if isinstance(val, list) and val:
            v = val[0]
            if isinstance(v, dict):
                return v.get("Description") or v.get("Name") or v.get("Text") or ""
            return str(v)
        if isinstance(val, dict):
            return val.get("Description") or val.get("Name") or val.get("Text") or ""
        if val is None:
            return ""
        return str(val)

    def _team_name(block):
        if not isinstance(block, dict):
            return ""
        for k in ("ShortClubName", "TeamName", "Name", "Abbreviation"):
            v = block.get(k)
            if v:
                desc = _first_desc(v)
                if desc:
                    return desc.strip()
        return ""

    def _score(block, primary_key, fallback_key):
        try:
            if isinstance(block, dict):
                v = block.get(primary_key)
                if v is None:
                    v = block.get(fallback_key)
                if v is None and "Score" in block:
                    v = block.get("Score")
                if v is None:
                    return None
                return int(v)
        except Exception:
            pass
        return None

    def _store_pair(team_a, team_b, score_a, score_b, status):
        _n0 = _norm_team(team_a)
        _n1 = _norm_team(team_b)
        _out[(_team_key(_n0), _team_key(_n1))] = {"home": score_a, "away": score_b, "status": status}
        _out[(_team_key(_n1), _team_key(_n0))] = {"home": score_b, "away": score_a, "status": status}

    # FIFA live feed first
    for _url in _fifa_urls:
        try:
            _resp = req.get(_url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
            if _resp.status_code != 200:
                continue

            for _ev in (_resp.json().get("Results") or []):
                _h = _team_name(_ev.get("Home"))
                _a = _team_name(_ev.get("Away"))
                if not _h or not _a:
                    continue

                _status = _status_text(
                    _ev.get("Status"),
                    _ev.get("MatchStatus"),
                    _ev.get("status"),
                    _ev.get("Phase"),
                    _ev.get("Progress"),
                    _ev.get("State"),
                    _ev.get("GameStatus"),
                    _ev.get("MatchState"),
                    _ev.get("CompetitionStatus"),
                    _ev.get("StatusDescription"),
                    _ev.get("Description"),
                )

                hs = _score(_ev.get("Home"), "Score", "HomeTeamScore")
                as_ = _score(_ev.get("Away"), "Score", "AwayTeamScore")
                if hs is None:
                    hs = _score(_ev, "HomeTeamScore", "HomeScore")
                if as_ is None:
                    as_ = _score(_ev, "AwayTeamScore", "AwayScore")

                if hs is not None and as_ is not None and _is_liveish(_status) and not _is_finishedish(_status):
                    _store_pair(_h, _a, hs, as_, _status or "live")

            if _out:
                break
        except Exception:
            continue

    # ESPN fallback
    if not _out:
        _espn_slugs = [
            "fifa.worldcup",
            "fifa.world",
            "global.2026-fifa-world-cup",
            "fifa.worldcup.2026",
        ]
        for _day in _days:
            for _slug in _espn_slugs:
                try:
                    _url = f"https://site.api.espn.com/apis/site/v2/sports/soccer/{_slug}/scoreboard?dates={_day}"
                    _resp = req.get(_url, timeout=8, headers={"User-Agent": "Mozilla/5.0"})
                    if _resp.status_code != 200:
                        continue
                    _data = _resp.json()
                    for _ev in _data.get("events", []):
                        _comp = (_ev.get("competitions") or [{}])[0]
                        _status = _comp.get("status", {}).get("type", {}).get("state", "").lower()

                        if _status not in ("in", "in_progress", "live", "active"):
                            continue

                        _teams = _comp.get("competitors", [])
                        if len(_teams) < 2:
                            continue

                        try:
                            _names = [t.get("team", {}).get("displayName", "") for t in _teams]
                            _scores = [int(t.get("score", 0) or 0) for t in _teams]
                        except Exception:
                            continue

                        _store_pair(_names[0], _names[1], _scores[0], _scores[1], _status)
                    if _out:
                        break
                except Exception:
                    continue
            if _out:
                break

    # Scoreaxis fallback
    if not _out:
        for _day in _iso_days:
            try:
                _url = f"https://www.scoreaxis.com/api/live-events?sport=1&date={_day}&timeZone=0"
                _resp = req.get(_url, timeout=8, headers={"User-Agent": "Mozilla/5.0"})
                if _resp.status_code != 200:
                    continue
                _events = (_resp.json().get("data", {}).get("events") or [])
                for _ev in _events:
                    _status = str(
                        _ev.get("statusText")
                        or _ev.get("status")
                        or _ev.get("matchStatus")
                        or _ev.get("phase")
                        or ""
                    ).casefold()

                    if _ev.get("statusFinished") is True:
                        continue
                    if not (_is_liveish(_status) or "live" in _status or "in progress" in _status or "half" in _status):
                        continue

                    _h = _ESPN_MAP.get((_ev.get("homeTeamName") or "").lower(), _ev.get("homeTeamName", ""))
                    _a = _ESPN_MAP.get((_ev.get("awayTeamName") or "").lower(), _ev.get("awayTeamName", ""))
                    if not _h or not _a:
                        continue

                    try:
                        _hs = int(_ev.get("homeScore", 0) or 0)
                        _as = int(_ev.get("awayScore", 0) or 0)
                    except Exception:
                        continue

                    _store_pair(_h, _a, _hs, _as, _status or "live")
                if _out:
                    break
            except Exception:
                continue

    return _out


def sync_results_to_excel(file_path, api_results, api_scores=None, live_scores=None, sheet_name="Sheet1"):
    """
    Write API results (and scores) back into the Excel Result / Score columns.
    Only fills blank cells so manual values stay intact.
    Skips any match that is currently live or in the live window.
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    headers = {
        str(ws.cell(1, c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value is not None
    }

    c_team1  = headers.get("Team 1")
    c_team2  = headers.get("Team 2")
    c_result = headers.get("Result")
    c_date   = headers.get("Date (NZDT)")
    c_time   = headers.get("Time (NZDT)")

    # Create Score column if it doesn't exist yet
    c_score = headers.get("Score")
    if not c_score:
        c_score = ws.max_column + 1
        ws.cell(1, c_score).value = "Score"

    if not c_team1 or not c_team2 or not c_result:
        raise ValueError("Missing required columns: Team 1, Team 2, Result")

    _now_sync = datetime.now(NZ_TZ)
    updated = 0

    for r in range(2, ws.max_row + 1):
        team1 = ws.cell(r, c_team1).value
        team2 = ws.cell(r, c_team2).value
        current_result = ws.cell(r, c_result).value

        if team1 is None or team2 is None:
            continue

        # Keep manual result if already present.
        if current_result is not None and str(current_result).strip() != "":
            # But still try to backfill a missing score for already-finished matches
            current_score = ws.cell(r, c_score).value
            if (current_score is None or str(current_score).strip() == "") and api_scores:
                k1, k2 = _team_key(team1), _team_key(team2)
                sc = api_scores.get((k1, k2)) or api_scores.get((k2, k1))
                if sc:
                    ws.cell(r, c_score).value = f"{sc[0]}-{sc[1]}"
                    updated += 1
            continue

        k1 = _team_key(team1)
        k2 = _team_key(team2)

        # Skip matches that are currently live.
        if live_scores and (live_scores.get((k1, k2)) or live_scores.get((k2, k1))):
            continue

        # Skip matches that are within the live window (0-130 min after kick-off)
        # to avoid writing a 0-0 placeholder as a final result.
        try:
            if c_date and c_time:
                _raw_date = ws.cell(r, c_date).value
                _raw_time = ws.cell(r, c_time).value
                if _raw_date is not None:
                    _match_dt = pd.to_datetime(_raw_date)
                    _t = parse_time(_raw_time)
                    if _t:
                        _match_dt = datetime.combine(_match_dt.date(), _t).replace(tzinfo=NZ_TZ)
                    else:
                        _match_dt = _match_dt.replace(tzinfo=NZ_TZ)
                    _secs = (_now_sync - _match_dt).total_seconds()
                    if 0 <= _secs <= (130 * 60):
                        continue  # still in live window
        except Exception:
            pass

        winner = api_results.get((k1, k2)) or api_results.get((k2, k1))
        if winner:
            ws.cell(r, c_result).value = winner
            # Also persist the score
            if api_scores:
                sc = api_scores.get((k1, k2)) or api_scores.get((k2, k1))
                if sc:
                    ws.cell(r, c_score).value = f"{sc[0]}-{sc[1]}"
            updated += 1

    if updated:
        wb.save(file_path)

    return updated


# ── Resolve which Excel file to use ─────────────────────────────────────────
# Priority: 1) temp-dir copy downloaded from SharePoint
#           2) local copy committed alongside app.py in the repo
LOCAL_FILE_PATH = str(BASE_DIR / "World Cup 2026 Comp.xlsx")

def _resolve_file_path():
    """Return the path to the Excel file, downloading from SharePoint if needed."""
    # Already have a fresh temp copy — use it
    if os.path.exists(FILE_PATH):
        return FILE_PATH

    # Try to download from SharePoint
    token = os.getenv("GRAPH_ACCESS_TOKEN")
    if token:
        try:
            download_workbook()
            if os.path.exists(FILE_PATH):
                return FILE_PATH
        except Exception as _dl_err:
            st.warning(f"⚠️ SharePoint download failed: {_dl_err}. Falling back to local file.")

    # Fall back to the repo copy
    if os.path.exists(LOCAL_FILE_PATH):
        return LOCAL_FILE_PATH

    st.error(
        "❌ Could not find the Excel workbook. "
        "Make sure `World Cup 2026 Comp.xlsx` is committed to your repo "
        "or that `GRAPH_ACCESS_TOKEN` is set in Streamlit secrets."
    )
    st.stop()

FILE_PATH = _resolve_file_path()

# mtime cache key → re-reads Excel only when file is saved
try:
    _mtime = os.path.getmtime(FILE_PATH)
except Exception:
    _mtime = 0

_api_results, _api_scores, _api_datetimes = fetch_api_results()
_live_scores    = fetch_live_scores()
_api_schedule   = fetch_schedule()
_goalscorers    = fetch_match_goalscorers()
_tourney_stats  = fetch_player_stats()
try:
    _ = sync_results_to_excel(FILE_PATH, _api_results, api_scores=_api_scores, live_scores=_live_scores)
except Exception as e:
    st.warning(f"Could not sync API results back to Excel: {e}")

try:
    _mtime = os.path.getmtime(FILE_PATH)
except Exception:
    _mtime = 0

df = load_data(_mtime, _path=FILE_PATH)

# Create an in-memory copy of the spreadsheet and fill missing Result/Score cells
# from API results so the UI updates immediately without requiring Excel writes.
# API dates are also applied to ALL rows to fix wrong Excel dates.
try:
    _df_copy = df.copy()
    _now_patch = datetime.now(NZ_TZ)
    for idx, _row in _df_copy.iterrows():
        t1 = _row.get("Team 1")
        t2 = _row.get("Team 2")
        if pd.isna(t1) or pd.isna(t2):
            continue
        k1 = _team_key(t1)
        k2 = _team_key(t2)

        # ── Always patch datetime from API for every row (fixes wrong Excel dates) ──
        # _api_schedule covers all matches (upcoming + finished); _api_datetimes is finished-only fallback
        _sched_dt = (
            _api_schedule.get((k1, k2)) or _api_schedule.get((k2, k1))
            if '_api_schedule' in globals() else None
        ) or (
            _api_datetimes.get((k1, k2)) or _api_datetimes.get((k2, k1))
            if '_api_datetimes' in globals() else None
        )
        if _sched_dt is not None:
            _df_copy.at[idx, "DateTime"]    = _sched_dt
            _df_copy.at[idx, "Date (NZDT)"] = pd.Timestamp(_sched_dt.date())

        # ── Only patch result/score for blank rows that have already kicked off ──
        cur = _row.get("Result")
        if pd.notna(cur) and str(cur).strip() != "":
            continue
        _match_dt = _df_copy.at[idx, "DateTime"]   # use the potentially-updated value
        if _match_dt is not None and pd.notna(_match_dt) and _match_dt > _now_patch:
            continue
        winner = _api_results.get((k1, k2)) or _api_results.get((k2, k1))
        if winner:
            _df_copy.at[idx, "Result"] = winner
        cur_score = _row.get("Score")
        if (pd.isna(cur_score) if isinstance(cur_score, float) else not cur_score):
            sc = _api_scores.get((k1, k2)) if '_api_scores' in globals() else None
            if sc:
                _df_copy.at[idx, "Score"] = f"{sc[0]}-{sc[1]}"
    df = _df_copy
except Exception:
    pass


# ── Smart refresh: keep polling while any today's match is live or recently finished ──────────
_now = datetime.now(NZ_TZ)
_today_matches = df[df["Date (NZDT)"].dt.date == _now.date()]
_has_pending = False
_pending_minutes = []

for _, _tm in _today_matches.iterrows():
    _r_val = str(_tm["Result"]).strip() if pd.notna(_tm["Result"]) else None
    _dt = _tm.get("DateTime")
    if _dt is None:
        continue
    _secs = (_now - _dt).total_seconds()
    _mins = _secs / 60.0

    # A match is "active" if it kicked off and hasn't been clearly finished
    # for more than 30 minutes (covers 90-min + 30-min buffer = 120 min).
    # We keep refreshing during this whole window regardless of whether the
    # API already has a result, so cache expiry is caught quickly.
    _in_window = 0 <= _secs <= (120 * 60)

    if _dt <= _now and _in_window:
        _has_pending = True
        _pending_minutes.append(_mins)
    elif not _r_val and _dt <= _now:
        # Past the window but still no result — keep polling slowly
        _t1 = str(_tm["Team 1"]).strip() if pd.notna(_tm["Team 1"]) else ""
        _t2 = str(_tm["Team 2"]).strip() if pd.notna(_tm["Team 2"]) else ""
        if not (_api_results.get((_team_key(_t1), _team_key(_t2))) or _api_results.get((_team_key(_t2), _team_key(_t1)))):
            _has_pending = True
            _pending_minutes.append(_mins)

if _has_pending:
    # During / just after a match: refresh every 45 s (matches fetch_api_results ttl).
    # Waiting > 30 min with no result: back off to every 3 minutes.
    if _pending_minutes and min(_pending_minutes) < 30:
        _delay_ms = 45_000    # 45 seconds — aligned with API cache TTL
    elif _pending_minutes and min(_pending_minutes) < 150:
        _delay_ms = 90_000    # 90 seconds for recently-finished matches
    else:
        _delay_ms = 180_000   # 3 minutes — slow poll for stale pending

    components.html(
        f"<script>setTimeout(()=>window.parent.location.reload(true),{_delay_ms});</script>",
        height=0
    )

# =========================
# ✅ COMPREHENSIVE FLAG MAP  (all WC2026 nations + extras)
# =========================
TEAM_FLAG_MAP = {
    # CONCACAF
    "USA": "us", "United States": "us", "Mexico": "mx", "Canada": "ca",
    "Jamaica": "jm", "Haiti": "ht", "Honduras": "hn", "Costa Rica": "cr",
    "Panama": "pa", "Trinidad and Tobago": "tt", "Guatemala": "gt",
    "El Salvador": "sv", "Cuba": "cu", "Nicaragua": "ni",
    # CONMEBOL
    "Argentina": "ar", "Brazil": "br", "Colombia": "co", "Ecuador": "ec",
    "Uruguay": "uy", "Chile": "cl", "Paraguay": "py", "Venezuela": "ve",
    "Bolivia": "bo", "Peru": "pe",
    # UEFA
    "Spain": "es", "England": "gb-eng", "France": "fr", "Germany": "de",
    "Portugal": "pt", "Italy": "it", "Netherlands": "nl", "Belgium": "be",
    "Switzerland": "ch", "Croatia": "hr", "Denmark": "dk", "Austria": "at",
    "Sweden": "se", "Norway": "no", "Scotland": "gb-sct", "Ukraine": "ua",
    "Türkiye": "tr", "Turkey": "tr", "Poland": "pl", "Wales": "gb-wls",
    "Greece": "gr", "Hungary": "hu", "Albania": "al", "Slovakia": "sk",
    "Serbia": "rs", "Czech Republic": "cz", "Finland": "fi", "Romania": "ro",
    "Slovenia": "si", "Iceland": "is", "North Macedonia": "mk", "Georgia": "ge",
    "Armenia": "am", "Bosnia and Herzegovina": "ba", "Bulgaria": "bg",
    "Estonia": "ee", "Latvia": "lv", "Lithuania": "lt", "Luxembourg": "lu",
    "Moldova": "md", "Montenegro": "me", "Northern Ireland": "gb-nir",
    "Republic of Ireland": "ie", "Ireland": "ie", "Kosovo": "xk",
    "Israel": "il", "Kazakhstan": "kz", "Belarus": "by", "Cyprus": "cy",
    "Faroe Islands": "fo", "Gibraltar": "gi", "Liechtenstein": "li",
    "Malta": "mt", "San Marino": "sm",
    # CAF
    "Morocco": "ma", "Egypt": "eg", "Nigeria": "ng", "Senegal": "sn",
    "Cameroon": "cm", "Ghana": "gh", "Ivory Coast": "ci",
    "Côte d'Ivoire": "ci", "Mali": "ml", "DR Congo": "cd",
    "Congo DR": "cd", "South Africa": "za", "Algeria": "dz",
    "Tunisia": "tn", "Zambia": "zm", "Zimbabwe": "zw", "Kenya": "ke",
    "Uganda": "ug", "Tanzania": "tz", "Angola": "ao", "Burkina Faso": "bf",
    "Benin": "bj", "Cape Verde": "cv", "Comoros": "km", "Ethiopia": "et",
    "Gabon": "ga", "Guinea": "gn", "Guinea-Bissau": "gw", "Libya": "ly",
    "Madagascar": "mg", "Malawi": "mw", "Mauritania": "mr",
    "Mozambique": "mz", "Namibia": "na", "Niger": "ne", "Rwanda": "rw",
    "Sudan": "sd", "Togo": "tg", "Liberia": "lr", "Sierra Leone": "sl",
    "Equatorial Guinea": "gq",
    # AFC
    "Japan": "jp", "South Korea": "kr", "Korea Republic": "kr",
    "Iran": "ir", "Iraq": "iq", "Jordan": "jo", "Uzbekistan": "uz",
    "Saudi Arabia": "sa", "Australia": "au", "Indonesia": "id",
    "Qatar": "qa", "UAE": "ae", "United Arab Emirates": "ae",
    "China": "cn", "Thailand": "th", "Vietnam": "vn", "Philippines": "ph",
    "Bahrain": "bh", "Kuwait": "kw", "Oman": "om", "Syria": "sy",
    "Kyrgyzstan": "kg", "Tajikistan": "tj", "India": "in", "Myanmar": "mm",
    "Palestine": "ps", "Lebanon": "lb", "Nepal": "np",
    # OFC
    "New Zealand": "nz", "Fiji": "fj", "Papua New Guinea": "pg",
    "Solomon Islands": "sb", "Vanuatu": "vu",
}

def get_flag_url(team):
    code = TEAM_FLAG_MAP.get(str(team).strip())
    return f"https://flagcdn.com/w80/{code}.png" if code else None

def flag_html(url):
    if url:
        return f'<img src="{url}" class="flag-img" onerror="this.style.display=\'none\'">'
    return '<span class="flag-blank"></span>'

# =========================
# ✅ LEADERBOARD
# =========================

completed_matches = int(df["Result"].notna().sum())
valid_matches = df[df["Team 1"].notna() & df["Team 2"].notna()]
total_matches = len(valid_matches)
completed_matches = valid_matches["Result"].notna().sum()
remaining_matches = total_matches - completed_matches
player_count = len([c for c in df.columns if "Pick" in c])
st.markdown(f"""
<div class="stat-row">
    <div class="stat-card" data-icon="🏟️">
        <div class="stat-label">Total Matches</div>
        <div class="stat-value blue">{total_matches}</div>
        <div class="stat-sublabel">Tournament Fixtures</div>
    </div>
    <div class="stat-card" data-icon="✅">
        <div class="stat-label">Completed</div>
        <div class="stat-value green">{completed_matches}</div>
        <div class="stat-sublabel">Results Available</div>
    </div>
    <div class="stat-card" data-icon="⏳">
        <div class="stat-label">Remaining</div>
        <div class="stat-value orange">{remaining_matches}</div>
        <div class="stat-sublabel">To Be Played</div>
    </div>
    <div class="stat-card" data-icon="👥">
        <div class="stat-label">Players</div>
        <div class="stat-value gold">{player_count}</div>
        <div class="stat-sublabel">Prediction Participants</div>
    </div>
</div>
""", unsafe_allow_html=True)

# ── Tournament Stats panel — right here, most visible spot ──────────────────
if _tourney_stats:
    _has_any = any(_tourney_stats.get(k) for k in ("top_scorers","top_assists","yellow_cards","red_cards"))
    if not _has_any:
        with st.expander("🔍 Stats Debug — click to see why data is missing", expanded=False):
            st.write(f"_tourney_stats keys: {list(_tourney_stats.keys())}")
            for k in ("top_scorers","top_assists","yellow_cards","red_cards"):
                st.write(f"  {k}: {len(_tourney_stats.get(k,[]))} items")
    def _ts_rows(items, val_cls="goals"):
        medals  = {1:"🥇", 2:"🥈", 3:"🥉"}
        p_cls   = {1:"gold", 2:"silver", 3:"bronze"}
        rows, prev, rank = "", None, 0
        for i, p in enumerate(items[:5], 1):
            if p.get("count") != prev: rank = i
            prev = p.get("count")
            _fc     = p.get("flag_code","")
            _flag   = f'<img src="https://flagcdn.com/w40/{_fc}.png" style="width:22px;height:15px;object-fit:cover;border-radius:2px;" onerror="this.style.display=\'none\'">' if _fc else ""
            _photo  = p.get("photo","")
            _medal  = medals.get(rank, f'<span style="color:#4dabf7;font-family:Bebas Neue,Impact,sans-serif;font-size:15px">{rank}</span>')
            _pcls   = p_cls.get(rank,"")
            rows += f"""
            <div class="ts-row">
                <div class="ts-pos {_pcls}">{_medal}</div>
                <div class="ts-player-wrap">
                    <div class="ts-photo-hover">
                        {"<img src='" + _photo + "' class='ts-photo-img' onerror=\"this.style.display='none'\">" if _photo else ""}
                    </div>
                    <div class="ts-flag-sm">{_flag}</div>
                    <div class="ts-info">
                        <div class="ts-name-sm">{p.get("name","")}</div>
                        <div class="ts-team-sm">{p.get("team","")}</div>
                    </div>
                </div>
                <div class="ts-val {val_cls}">{p.get("count",0)}</div>
            </div>"""
        return rows or '<div style="color:#506080;font-size:12px;padding:10px 4px">No data yet</div>'

    _s_html = _ts_rows(_tourney_stats.get("top_scorers",[]),  "goals")
    _a_html = _ts_rows(_tourney_stats.get("top_assists",[]),  "assists")
    _y_html = _ts_rows(_tourney_stats.get("yellow_cards",[]), "yellow")
    _r_html = _ts_rows(_tourney_stats.get("red_cards",[]),    "red")
    st.markdown(f"""
    <div class="ts-panel">
        <div class="ts-card"><div class="ts-card-title">⚽ TOP SCORERS</div>{_s_html}</div>
        <div class="ts-card"><div class="ts-card-title">🎯 TOP ASSISTS</div>{_a_html}</div>
        <div class="ts-card"><div class="ts-card-title">🟨 YELLOW CARDS</div>{_y_html}</div>
        <div class="ts-card"><div class="ts-card-title">🟥 RED CARDS</div>{_r_html}</div>
    </div>
    """, unsafe_allow_html=True)

pick_cols = [col for col in df.columns if "Pick" in col]

records = []
for col in pick_cols:
    player_name = col.replace(" Pick", "")
    for _, row in df.iterrows():
        t1 = str(row["Team 1"]) if pd.notna(row["Team 1"]) else "TBD"
        t2 = str(row["Team 2"]) if pd.notna(row["Team 2"]) else "TBD"
        records.append({
            "Player": player_name,
            "Match":  f"{t1} vs {t2}",
            "Pick":   row[col],
            "Result": row["Result"],
        })

data = pd.DataFrame(records)
data["Correct"] = (data["Pick"] == data["Result"]).astype("float")

played = data[data["Result"].notna()]

leaderboard = (
    played.groupby("Player")
    .agg(Points=("Correct", "sum"), Matches=("Correct", "count"))
    .reset_index()
    .sort_values("Points", ascending=False)
    .reset_index(drop=True)
)

# Dense rank with clean symbols
def compute_ranks(points_list):
    from collections import Counter
    count = Counter(points_list)
    sorted_uniq = sorted(set(points_list), reverse=True)
    pts_to_dense = {pts: i + 1 for i, pts in enumerate(sorted_uniq)}
    podium = {1: "🥇", 2: "🥈", 3: "🥉"}

    result = []
    for pts in points_list:
        dr = pts_to_dense[pts]
        # Top 3 get medals, everyone else gets a plain dense rank number.
        # Ties share the same number automatically; no '=4' / '=5' symbols.
        result.append(podium.get(dr, str(dr)))
    return result

leaderboard.insert(0, "Rank", compute_ranks(leaderboard["Points"].tolist()))
leaderboard["Accuracy"] = (leaderboard["Points"] / leaderboard["Matches"] * 100).round(1).astype(str) + "%"

st.markdown('<div class="section-title">🏆 Leaderboard</div>', unsafe_allow_html=True)


_PALETTE = [
    "#ff6b6b","#ffd43b","#69db7c","#4dabf7","#ff922b",
    "#cc5de8","#20c997","#f06595","#74c0fc","#a9e34b",
    "#ff8787","#63e6be","#ffa94d","#b2f2bb","#da77f2",
    "#ff6eb4","#94d82d","#66d9e8","#ffc078","#a5d8ff",
]
_sorted_p = sorted(leaderboard["Player"].unique())
_cmap     = {p: _PALETTE[i % len(_PALETTE)] for i, p in enumerate(_sorted_p)}

_max_pts = leaderboard["Points"].max() if len(leaderboard) > 0 else 1
# Bar width = accuracy % (points/matches) so tied players show correctly
# e.g. 3/6 = 50%, 3/4 = 75% – visually meaningful

def _initials(name):
    parts = str(name).strip().split()
    return (parts[0][0] + (parts[-1][0] if len(parts) > 1 else parts[0][-1])).upper()

def _num_cls(rs):
    return "p1" if rs in ("🥇","🏅") else "p2" if rs=="🥈" else "p3" if rs=="🥉" else "pn"

def _row_cls(rs):
    return "rank-1" if rs in ("🥇","🏅") else "rank-2" if rs=="🥈" else "rank-3" if rs=="🥉" else ""

def _make_lb_row(_pl, _pts, _mat, _rk, _acc, _col):
    _pct = round((_pts / _mat) * 100) if _mat > 0 else 0
    _av  = _initials(_pl)
    _nc  = _num_cls(_rk)
    _rc  = _row_cls(_rk)
    return (f'<div class="lb-row {_rc}" style="--pc:{_col}">'
            f'<div class="lb-num {_nc}">{_rk}</div>'
            f'<div class="lb-avatar" style="background:{_col}">{_av}</div>'
            f'<div class="lb-name">{_pl}</div>'
            f'<div class="lb-bar-wrap"><div class="lb-bar" style="background:{_col};width:{_pct}%"></div></div>'
            f'<div class="lb-score">{_pts}<sup>/{_mat}</sup></div>'
            f'<div class="lb-acc">{_acc}</div>'
            f'</div>')

_lb_records = [(str(r["Player"]), int(r["Points"]), int(r["Matches"]),
                str(r["Rank"]), str(r["Accuracy"]), _cmap.get(str(r["Player"]), "#4dabf7"))
               for _, r in leaderboard.iterrows()]
_half      = (len(_lb_records) + 1) // 2
_left_html  = "".join(_make_lb_row(*rec) for rec in _lb_records[:_half])
_right_html = "".join(_make_lb_row(*rec) for rec in _lb_records[_half:])

st.markdown(
    f'<div class="lb-container">'
    f'<div class="lb-grid">'
    f'<div class="lb-col">{_left_html}</div>'
    f'<div class="lb-col">{_right_html}</div>'
    f'</div></div>',
    unsafe_allow_html=True
)

# =========================
# 📅 MATCH SCHEDULE
# =========================
st.markdown("""
<div class="pitch-divider">
 </div>
""", unsafe_allow_html=True)




now = datetime.now(NZ_TZ)
today = now.date()
selected_date = st.date_input("Select Date", value=today)

matches = df[df["Date (NZDT)"].dt.date == selected_date]

# ── Filter out non-WC rows (e.g. domestic league games from API) ─────────────
_wc_teams = set(t.casefold() for t in TEAM_FLAG_MAP.keys())

def _is_wc_team(name):
    """True if team name is a known WC2026 nation."""
    if not name or str(name).strip().casefold() in ("tbd","tba","nan",""):
        return True   # TBD slots are valid — keep them
    return str(name).strip().casefold() in _wc_teams or get_flag_url(str(name).strip()) is not None

matches = matches[matches.apply(
    lambda r: _is_wc_team(str(r["Team 1"])) and _is_wc_team(str(r["Team 2"])), axis=1
)]

if matches.empty:
    st.info("⚽ No matches scheduled for this date.")
else:
    for _, row in matches.iterrows():
        team1  = str(row["Team 1"]) if pd.notna(row["Team 1"]) else "TBD"
        team2  = str(row["Team 2"]) if pd.notna(row["Team 2"]) else "TBD"
        dt     = row["DateTime"]
        result = str(row["Result"]).strip() if pd.notna(row["Result"]) else None
        k1 = _team_key(team1)
        k2 = _team_key(team2)

        live = (_live_scores.get((k1, k2)) or _live_scores.get((k2, k1))) if '_live_scores' in globals() else None
        api_result = _api_results.get((k1, k2)) or _api_results.get((k2, k1))
        api_score  = (_api_scores.get((k1, k2)) or _api_scores.get((k2, k1))) if '_api_scores' in globals() else None

        # Override dt from API schedule (covers upcoming + finished) — Excel dates never used
        _sched_dt = (
            (_api_schedule.get((k1, k2)) or _api_schedule.get((k2, k1)))
            if '_api_schedule' in globals() else None
        ) or (
            (_api_datetimes.get((k1, k2)) or _api_datetimes.get((k2, k1)))
            if '_api_datetimes' in globals() else None
        )
        if _sched_dt is not None:
            dt = _sched_dt

        # API is authoritative — always prefer API result + score over Excel
        if api_result:
            result = str(api_result).strip()
        if api_score is None:
            # Fall back to persisted score from Excel only if API has nothing
            _raw_score = row.get("Score") if hasattr(row, "get") else None
            if _raw_score and pd.notna(_raw_score):
                try:
                    _parts = str(_raw_score).split("-")
                    if len(_parts) == 2:
                        api_score = (int(_parts[0]), int(_parts[1]))
                except Exception:
                    pass
        has_result = result is not None and str(result).strip() != ""
        live_state = str((live or {}).get("status", "")).casefold()

        # ── Classify feed status ──────────────────────────────────────────────
        # is_finished_feed must be evaluated first — used inside is_live_feed.
        is_finished_feed = bool(live and any(k in live_state for k in (
            "finished", "completed", "final", "post", "postperiod", "ft", "aet"
        )))

        # Only treat as live if the feed is NOT already finished.
        is_live_feed = bool(
            live and not is_finished_feed and (
                any(k in live_state for k in (
                    "live", "in progress", "inprogress", "1st half", "2nd half",
                    "half time", "halftime", "ht", "extra time", "penalty"
                ))
                or live_state in ("in", "active", "in_progress", "inprogress")
            )
        )

        # Time-based live-window guard (0 – 130 min after kick-off).
        # Prevents a stale 0-0 "Draw" from the API being shown as final
        # while the match is still in progress or only just finished.
        _secs_since_ko = (now - dt).total_seconds() if dt else 9999
        _in_live_window = dt is not None and 0 <= _secs_since_ko <= (130 * 60)

        if is_live_feed:
            # Live data is available – clear any stale Excel / API result.
            has_result = False
            result = None
        elif _in_live_window and has_result and _is_draw_result(result):
            # A "Draw" within the live window is likely a pre-match / in-progress
            # 0-0 placeholder written to Excel by a previous run.  Treat as pending.
            # BUT if the API explicitly confirms it as finished, trust the API.
            api_confirmed = bool(api_result)
            if not api_confirmed:
                has_result = False
                result = None
        elif not has_result and api_result and (dt is None or dt <= now):
            # API has a result — trust it, but only for matches that have kicked off.
            result = str(api_result).strip()
            has_result = True

        # If a live feed has clearly finished and it's a draw, show Draw.
        finalized_live_draw = bool(
            is_finished_feed
            and live
            and live.get("home") is not None
            and live.get("away") is not None
            and live.get("home") == live.get("away")
        )
        if finalized_live_draw and not has_result:
            result = "Draw"
            has_result = True

        f1 = flag_html(get_flag_url(team1))
        f2 = flag_html(get_flag_url(team2))

        # Status: live feed first, then finished, then scheduled/pending.
        if is_live_feed:
            status = "LIVE"
        elif has_result or finalized_live_draw:
            status = "Finished"
        elif dt:
            diff = (now - dt).total_seconds()
            if diff < 0:
                status = "Upcoming"
            else:
                status = "Result Pending"
        else:
            status = "TBC"

        # ── Centre display ───────────────────────────────────────────────────
        t1_cls = t2_cls = ""
        if status == "Finished":
            # Try to get the score from API score dict or live feed
            _fin_score = api_score or (
                (live.get("home"), live.get("away")) if live else None
            )
            if result and _is_draw_result(result):
                if _fin_score and _fin_score[0] is not None:
                    score_html = (
                        f'<span class="result-draw">'
                        f'{_fin_score[0]} — {_fin_score[1]}'
                        f'<br><small style="font-size:11px;letter-spacing:2px;opacity:0.8;">DRAW</small>'
                        f'</span>'
                    )
                else:
                    score_html = '<span class="result-draw">⚖ DRAW</span>'
            elif result and _team_key(result) == _team_key(team1):
                t1_cls, t2_cls = "winner", "loser"
                if _fin_score and _fin_score[0] is not None:
                    score_html = (
                        f'<span class="result-win score-line">'
                        f'{_fin_score[0]} — {_fin_score[1]}'
                        f'<br><span style="font-size:13px;letter-spacing:2px;color:#ffd700;font-family:\'Bebas Neue\',Impact,sans-serif;">✓ {team1} WINS</span>'
                        f'</span>'
                    )
                else:
                    score_html = f'<span class="result-win">✓ {team1}<br>WINS</span>'
            elif result and _team_key(result) == _team_key(team2):
                t1_cls, t2_cls = "loser", "winner"
                if _fin_score and _fin_score[0] is not None:
                    score_html = (
                        f'<span class="result-win score-line">'
                        f'{_fin_score[0]} — {_fin_score[1]}'
                        f'<br><span style="font-size:13px;letter-spacing:2px;color:#ffd700;font-family:\'Bebas Neue\',Impact,sans-serif;">✓ {team2} WINS</span>'
                        f'</span>'
                    )
                else:
                    score_html = f'<span class="result-win">✓ {team2}<br>WINS</span>'
            else:
                score_html = f'<span class="result-win">{result}</span>'
        elif status == "LIVE":
            if live:
                hs = live.get("home", 0)
                as_ = live.get("away", 0)
                score_html = f'<span class="live-pulse">🔴 {hs} - {as_}</span>'
            else:
                score_html = '<span class="live-pulse">● LIVE</span>'
        else:
            score_html = '<span class="vs-text">VS</span>' if status == "Upcoming" else '<span class="result-win">WAITING<br>FOR RESULT</span>'

        time_str = dt.strftime("%d %b  ·  %I:%M %p NZT") if dt else "Time TBC"

        badge_map = {
            "LIVE":           '<span class="badge badge-live">🔴 LIVE</span>',
            "Upcoming":       '<span class="badge badge-upcoming">⏳ UPCOMING</span>',
            "Result Pending": '<span class="badge badge-finished">⏳ RESULT PENDING</span>',
            "Finished":       '<span class="badge badge-finished">✅ FINISHED</span>',
            "TBC":            '<span class="badge badge-tbc">🕐 TBC</span>',
        }
        if   status == "LIVE": card_cls = "match-card live"
        elif status in ("Finished", "Result Pending"): card_cls = "match-card finished"
        else: card_cls = "match-card"

        # ── Goalscorers — keyed by team_key, not home/away ──────────────────
        _gs_data = None
        if '_goalscorers' in globals():
            _gs_data = _goalscorers.get((k1, k2)) or _goalscorers.get((k2, k1))

        def _gs_col(goals, side="left"):
            cls = "gs-col" if side == "left" else "gs-col right"
            if not goals:
                return f'<div class="{cls}"></div>'
            items = ""
            for g in goals:
                if not g.get("name"): continue
                og = ' <span style="color:#ff7070;font-size:9px">(OG)</span>' if g.get("og") else ""
                items += (
                    f'<div class="gs-entry">'
                    f'<span class="gs-ball">⚽</span>'
                    f'<span class="gs-name">{g["name"]}{og}</span>'
                    f'<span class="gs-min">{g["minute"]}\'</span>'
                    f'</div>'
                )
            return f'<div class="{cls}">{items}</div>'

        _show_gs = bool(_gs_data and status in ("Finished", "LIVE"))
        if _show_gs and _gs_data:
            # _gs_data is {team_key: [goals]} — look up by each team's key directly
            _hg = _gs_data.get(k1, [])
            _ag = _gs_data.get(k2, [])
        else:
            _hg, _ag = [], []

        _gs_left  = _gs_col(_hg, "left")  if _show_gs else '<div class="gs-col"></div>'
        _gs_right = _gs_col(_ag, "right") if _show_gs else '<div class="gs-col right"></div>'

        st.markdown(f"""
        <div class="{card_cls}">
            <div class="match-inner">
                <div class="team-left">
                    {f1}
                    <span class="team-name {t1_cls}">{team1}</span>
                </div>
                {_gs_left}
                <div class="vs-center">
                    {score_html}
                </div>
                {_gs_right}
                <div class="team-right">
                    <span class="team-name {t2_cls}">{team2}</span>
                    {f2}
                </div>
                <div class="match-meta">
                    <div class="match-date">{time_str}</div>
                    {badge_map[status]}
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

# =========================
# 👥 PLAYER SUMMARY

# =========================
st.markdown('<div class="section-title">👥 Player Summary</div>', unsafe_allow_html=True)
st.dataframe(leaderboard, use_container_width=True, hide_index=True)

# =========================
# 🔍 PLAYER DETAIL
# =========================
st.markdown('<div class="section-title">🔍 Player Detail</div>', unsafe_allow_html=True)

player_sel = st.selectbox("Select Player", leaderboard["Player"].tolist())

p_df     = data[data["Player"] == player_sel].copy()
p_played = p_df[p_df["Result"].notna()]

correct  = int(p_played["Correct"].sum())
total    = p_played.shape[0]
accuracy = round((correct / total) * 100, 1) if total else 0

acc_cls      = "green" if accuracy >= 60 else "orange" if accuracy >= 35 else "red"
p_rank_s     = leaderboard.loc[leaderboard["Player"] == player_sel, "Rank"]
p_rank       = p_rank_s.iloc[0] if not p_rank_s.empty else "—"
acc_bar_w    = min(int(accuracy), 100)

st.markdown(f"""
<div class="stat-row">
    <div class="stat-card" data-icon="⚽">
        <div class="stat-label">Points</div>
        <div class="stat-value gold">{correct}</div>
        <div class="stat-sublabel">Correct Predictions</div>
    </div>
    <div class="stat-card" data-icon="📋">
        <div class="stat-label">Matches Played</div>
        <div class="stat-value blue">{total}</div>
        <div class="stat-sublabel">Results Available</div>
    </div>
    <div class="stat-card" data-icon="🎯">
        <div class="stat-label">Accuracy</div>
        <div class="stat-value {acc_cls}">{accuracy}%</div>
        <div class="accuracy-track"><div class="accuracy-fill" style="width:{acc_bar_w}%"></div></div>
    </div>
    <div class="stat-card" data-icon="🏆">
        <div class="stat-label">Leaderboard Rank</div>
        <div class="stat-value" style="font-size:42px;line-height:1.2;">{p_rank}</div>
        <div class="stat-sublabel">Current Standing</div>
    </div>
</div>
""", unsafe_allow_html=True)

st.dataframe(
    p_df[["Match", "Pick", "Result", "Correct"]],
    use_container_width=True,
    hide_index=True,
)

# =========================
# ⚠️ MISSING PICKS
# =========================
st.markdown('<div class="section-title">⚠️ Missing Picks</div>', unsafe_allow_html=True)

_vm          = df[df["Team 1"].notna() & df["Team 2"].notna()].reset_index(drop=True)
_gs_matches  = _vm.iloc[:72]
_ko_matches  = _vm.iloc[72:]
_ko_total    = len(_ko_matches)

# ── Round 1 ──
st.markdown("**Round 1 — Group Stage (72 matches)**")
_r1_missing = []
for col in pick_cols:
    p_name = col.replace(" Pick", "")
    filled = int(_gs_matches[col].notna().sum())
    miss   = 72 - filled
    if miss > 0:
        _r1_missing.append({"Player": p_name, "Filled": filled, "Missing": miss})
if _r1_missing:
    st.dataframe(pd.DataFrame(_r1_missing), use_container_width=True, hide_index=True)
else:
    st.success("✅ All players have completed their Round 1 picks!")

# ── Round of 32 ──
if _ko_total > 0:
    st.markdown(f"**Round of 32 — Knockout ({_ko_total} matches)**")
    _r2_missing = []
    for col in pick_cols:
        p_name = col.replace(" Pick", "")
        filled = int(_ko_matches[col].notna().sum())
        miss   = _ko_total - filled
        if miss > 0:
            _r2_missing.append({"Player": p_name, "Filled": filled, "Missing": miss})
    if _r2_missing:
        st.dataframe(pd.DataFrame(_r2_missing), use_container_width=True, hide_index=True)
    else:
        st.success("✅ All players have completed their Round of 32 picks!")
else:
    st.info("ℹ️ Round of 32 fixtures not yet available in the spreadsheet.")
